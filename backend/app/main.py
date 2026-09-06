from fastapi import FastAPI, Request, HTTPException, UploadFile, File, BackgroundTasks
from fastapi.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles
from fastapi.responses import StreamingResponse, FileResponse
from contextlib import asynccontextmanager
import asyncio
import json
import os
import threading
import time as _time
import uuid
import contextvars
from typing import AsyncGenerator, Dict, List, Optional

from tools.base import current_document_ids

from settings import APP_HOST, APP_PORT, DEBUG, MAX_UPLOAD_SIZE, UPLOAD_DIR, OUTPUTS_DIR
from database.models import init_db, get_db, shutdown_db
from database.crud import (
    create_conversation, get_conversation, get_all_conversations,
    add_message, get_conversation_messages, update_conversation_title,
    delete_conversation as db_delete_conversation,
    update_message, get_message, create_document, update_document_status, get_documents,
    delete_message as db_delete_message, delete_document as db_delete_document, get_document,
    get_all_agents, get_agent, get_agent_by_name, create_agent,
    update_agent, delete_agent,
    add_mcp_server, get_all_mcp_servers, get_enabled_mcp_servers,
    toggle_mcp_server, remove_mcp_server, update_mcp_server_disabled_tools,
    update_conversation_tags, update_conversation_agent,
    create_note, get_all_notes, delete_note, get_notes_for_conversation,
    get_message_versions, update_assistant_message_full
)
from mcp_client.client import MCPClientManager, MCPServerConfig
from tools.tool_executor import ToolExecutor, custom_tool_catalogue, custom_tool_names
from llm_client.client import LLMClient
from backend.settings import settings_manager
from backend.database.backup import backup_scheduler

# Initialize MCP client manager
mcp_manager = MCPClientManager()


async def _bootstrap_default_provider():
    """On first run, seed a default LLM provider from the legacy settings
    (llama_cpp_base_url) so existing installs keep working unchanged."""
    try:
        from database.provider_crud import list_providers, create_provider
        from tools.provider_service import fetch_models
        async with get_db() as db:
            providers = await list_providers(db)
            if providers:
                return
            base_url = settings_manager.get_settings().get("llama_cpp_base_url", "http://localhost:8080")
            try:
                models = await fetch_models(base_url, None, timeout=15)
                print(f"[PROVIDER] bootstrapped default provider with {len(models)} models")
            except Exception as e:
                models = []
                print(f"[PROVIDER] bootstrap fetch failed (models will be empty): {e}")
            await create_provider(
                db, "llama.cpp", base_url.rstrip("/"), api_key=None,
                models=models, is_default=1, enabled=1
            )
            await db.commit()
    except Exception as e:
        print(f"[PROVIDER] bootstrap failed: {e}")


async def _backfill_context_windows():
    """Re-fetch cached provider models that lack a context_window.

    The models JSON column was written before fetch_models learned to capture
    each model's window (llama.cpp reports --ctx-size in /v1/models). One cheap
    refresh per affected provider fixes existing installs without asking the
    user to re-add anything. Providers already carrying windows are skipped, so
    this is a no-op on normal restarts.
    """
    try:
        from database.provider_crud import list_providers, update_provider
        from tools.provider_service import fetch_models
        async with get_db() as db:
            providers = await list_providers(db, include_api_key=True)
            stale = [p for p in providers
                     if (p.get("models") or []) and not any(m.get("context_window") for m in p["models"])]
            for p in stale:
                try:
                    fresh = await fetch_models(p["base_url"], p.get("api_key"), timeout=8)
                    if any(m.get("context_window") for m in fresh):
                        await update_provider(db, p["id"], models=fresh)
                        print(f"[PROVIDER] {p['name']}: backfilled context windows for {len(fresh)} models")
                except Exception as e:
                    print(f"[PROVIDER] {p.get('name')}: context backfill skipped: {str(e)[:120]}")
            await db.commit()
    except Exception as e:
        print(f"[PROVIDER] context backfill failed: {e}")


async def _retire_default_agent():
    """Soft-delete the seeded 'Default' agent row, once.

    It duplicated the meaning of an empty agent selection (no agent → settings
    system prompt + all tools) and showed up as a second, confusing 'Default'
    entry next to the chat dropdown's own default option. Retiring it leaves the
    null selection as the single default path. Conversations still bound to it
    resolve to no agent, which is exactly what the row pretended to be.
    """
    try:
        from sqlalchemy import update as _upd
        from database.models import Agent as _Agent
        async with get_db() as db:
            result = await db.execute(
                _upd(_Agent)
                .where(_Agent.name == "Default")
                .where(_Agent.is_active == 1)
                .values(is_active=0)
            )
            if result.rowcount:
                await db.commit()
                print(f"[AGENT] retired {result.rowcount} seeded Default agent row(s)")
    except Exception as e:
        print(f"[AGENT] Default retirement failed: {e}")


async def _reconcile_pending_tool_messages():
    """On startup, mark any in-flight tool calls (approval/pending) as interrupted.

    If the app restarted while a turn was waiting for approval or mid-tool,
    the in-memory gate/progress is gone but the placeholder row is already
    committed (incremental persist). Without this, the UI would show a spinner
    forever. Marking as error makes the state explicit and lets the user retry.
    """
    try:
        from sqlalchemy import select as _sel
        from database.models import Message as _Msg
        async with get_db() as db:
            result = await db.execute(_sel(_Msg).where(_Msg.role == "assistant"))
            rows = result.scalars().all()
            fixed = 0
            for msg in rows:
                meta = msg.extra_metadata or {}
                blocks = meta.get("blocks") or []
                changed = False
                for b in blocks:
                    if b.get("type") == "tool_call" and b.get("status") in ("approval", "pending", "starting", "running"):
                        b["status"] = "error"
                        b["result"] = {"error": "Interrupted — app restarted or connection lost while waiting for tool/approval. Please retry by sending a follow-up message."}
                        b["progress"] = 0
                        changed = True
                if changed:
                    msg.extra_metadata = meta
                    # also keep tool_calls column consistent
                    tool_calls = []
                    for idx, b in enumerate([x for x in blocks if x.get("type") == "tool_call"]):
                        tid = b.get("id") or f"{b.get('name','tool')}_{idx}"
                        tool_calls.append({"id": tid, "name": b.get("name"), "arguments": b.get("arguments",{}), "status": b.get("status","error"), "progress": 0})
                    msg.tool_calls = tool_calls
                    fixed += 1
            if fixed:
                await db.commit()
                print(f"[RECONCILE] marked {fixed} interrupted assistant message(s) as error")
    except Exception as e:
        print(f"[RECONCILE] failed: {e}")


@asynccontextmanager
async def lifespan(app: FastAPI):
    # Startup
    await init_db()  # create_all for new tables + ALTER TABLE migrations
    await _bootstrap_default_provider()
    await _backfill_context_windows()
    await _retire_default_agent()
    await mcp_manager.initialize()
    backup_scheduler.start()
    yield
    # Shutdown
    await backup_scheduler.stop()
    await mcp_manager.cleanup()
    try:
        from tools.web_extract import _reset_session as _close_web_session
        await _close_web_session()
    except Exception:
        pass
    await shutdown_db()

app = FastAPI(title="LLM UI with MCP Support", lifespan=lifespan)

# Mount static files and templates
app.mount("/static", StaticFiles(directory="frontend/static"), name="static")
# Mount uploads directory so uploaded files are publicly accessible
os.makedirs(UPLOAD_DIR, exist_ok=True)
app.mount("/uploads", StaticFiles(directory=UPLOAD_DIR), name="uploads")
# Mount outputs directory (agent platform: job outputs, generated files)
app.mount("/outputs", StaticFiles(directory=OUTPUTS_DIR), name="outputs")
templates = Jinja2Templates(directory="frontend/templates")
# ══════════════════════════════ PWA ═══════════════════════════════
_FRONTEND_STATIC = os.path.join(
    os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
    "frontend", "static"
)

@app.get("/manifest.webmanifest", include_in_schema=False)
async def pwa_manifest():
    return FileResponse(
        os.path.join(_FRONTEND_STATIC, "manifest.webmanifest"),
        media_type="application/manifest+json",
    )

@app.get("/sw.js", include_in_schema=False)
async def pwa_service_worker():
    return FileResponse(
        os.path.join(_FRONTEND_STATIC, "sw.js"),
        media_type="text/javascript",
    )
# Initialize components
llm_client = LLMClient()
tool_executor = ToolExecutor(mcp_manager)

# Set TTS service in settings manager
settings_manager.set_tts_service(tool_executor.tts_service)

# Initialize STT service
from tools.stt_service import STTService, STTConfig
stt_service = STTService(STTConfig.from_settings(settings_manager.get_settings()))


@app.get("/")
async def index(request: Request):
    """Render main chat interface"""
    return templates.TemplateResponse(request, "index.html")


@app.get("/settings")
async def settings_page(request: Request):
    """Redirect to main page (settings are now a modal)"""
    return templates.TemplateResponse(request, "index.html")


@app.get("/knowledge")
async def knowledge_page(request: Request):
    """Redirect to main page (knowledge base is now a modal)"""
    return templates.TemplateResponse(request, "index.html")


@app.get("/agents")
async def agents_page(request: Request):
    """Redirect to main page (agents are now a modal)"""
    return templates.TemplateResponse(request, "index.html")


@app.get("/api/conversations")
async def list_conversations():
    """Get all conversations"""
    async with get_db() as db:
        conversations = await get_all_conversations(db)
        return {"conversations": conversations}


@app.post("/api/conversations")
async def new_conversation(request: Request):
    """Create a new conversation"""
    data = await request.json()
    title = data.get("title", "New Chat")
    agent_id = data.get("agent_id")
    
    async with get_db() as db:
        conversation = await create_conversation(db, title, agent_id)
        return {"conversation": conversation}


@app.get("/api/conversations/{conversation_id}")
async def get_conversation_detail(conversation_id: str):
    """Get conversation with messages"""
    async with get_db() as db:
        conversation = await get_conversation(db, conversation_id)
        if not conversation:
            raise HTTPException(status_code=404, detail="Conversation not found")
        
        messages = await get_conversation_messages(db, conversation_id)
        return {
            "conversation": conversation,
            "messages": messages
        }


@app.get("/api/conversations/{conversation_id}/context")
async def get_conversation_context(conversation_id: str):
    """Preview what the model will see on the next turn for this conversation.
    Recomputes system prompt + tools + counts without calling the LLM — so past
    chats show a useful preview instead of 'No context yet'."""
    async with get_db() as db:
        conversation = await get_conversation(db, conversation_id)
        if not conversation:
            raise HTTPException(status_code=404, detail="Conversation not found")
        agent_config = None
        if conversation.get("agent_id") is not None:
            agent = await get_agent(db, conversation["agent_id"])
            if agent:
                agent_config = {
                    "system_prompt": agent.system_prompt,
                    "model": agent.model,
                    "provider_id": getattr(agent, "provider_id", None),
                    "temperature": agent.temperature,
                    "top_k": agent.top_k,
                    "max_tokens": agent.max_tokens,
                    "enable_rag": bool(agent.enable_rag),
                    "enabled_tools": agent.enabled_tools or [],
                    "enabled_mcp_servers": agent.enabled_mcp_servers or [],
                    "enabled_skills": agent.enabled_skills or []
                }
        from datetime import datetime
        current_date = datetime.now().strftime("%Y-%m-%d")
        identity_hint = "You are running inside llm_ui — a local, single-user chat app on this machine. Conversations, memory, agents, skills and files you manage are all inside llm_ui unless the user explicitly names an external app. Memory: relevant durable facts may be appended to the user's message in a <relevant_memories> block — use them. Before answering about past work, preferences, projects or people, call memory_search. When the user states a durable fact or preference (or says 'remember X'), call memory_write immediately: concise, standalone, tagged. Never store secret values — at most where they are kept."
        system_prompt_content = agent_config["system_prompt"] if agent_config and agent_config.get("system_prompt") else ""
        if system_prompt_content:
            system_prompt_content = f"{system_prompt_content}\n\n{identity_hint}\n\nCurrent date: {current_date}"
        else:
            system_prompt_content = f"You are a helpful AI assistant. {identity_hint} Current date: {current_date}"
        try:
            from database.memory_crud import get_memory_for_injection
            memory_block = await get_memory_for_injection(
                db,
                agent_id=conversation.get("agent_id"),
                conversation_id=conversation_id
            )
            if memory_block:
                system_prompt_content += "\n\n" + memory_block
        except Exception as e:
            print(f"[MEMORY] context preview failed: {e}")
        try:
            from tools.skills_tool import skill_index
            idx = skill_index()
            if idx:
                enabled_skills = set()
                if agent_config and agent_config.get("enabled_skills"):
                    enabled_skills = set(agent_config["enabled_skills"])
                if enabled_skills:
                    idx = "\n".join(
                        line for line in idx.splitlines()
                        if line.startswith("- ")
                        and line[2:].split(":", 1)[0].strip() in enabled_skills
                    )
                if idx:
                    system_prompt_content += (
                        "\n\n### Available skills\n"
                        "Call load_skill(name) to load a skill's full instructions "
                        "when it is relevant to the current task.\n" + idx
                    )
        except Exception as e:
            print(f"[SKILLS] context preview failed: {e}")
        # messages + tools count preview — use same KV-aware replay as stream handler
        try:
            messages = await get_conversation_messages(db, conversation_id)
            llm_messages = []
            for msg in messages:
                role = msg["role"]
                content = msg["content"]
                metadata = msg.get("metadata") or {}
                if role == "user" and (metadata.get("files") or []):
                    # Keep preview simple: one entry per user message with files
                    llm_messages.append({"role": role, "content": content})
                    continue
                blocks = (metadata.get("blocks") or []) if isinstance(metadata, dict) else []
                tool_blocks = [b for b in blocks if b.get("type") == "tool_call"] if role == "assistant" else []
                has_tools = bool(role == "assistant" and msg.get("tool_calls") and tool_blocks)
                if not has_tools:
                    m = {"role": role, "content": content}
                    if role == "assistant" and msg.get("thinking"):
                        m["reasoning_content"] = msg["thinking"]
                    llm_messages.append(m)
                else:
                    first_tool_pos = next((i for i, b in enumerate(blocks) if b.get("type") == "tool_call"), len(blocks))
                    last_tool_pos = len(blocks) - 1 - next((i for i, b in enumerate(reversed(blocks)) if b.get("type") == "tool_call"), len(blocks))
                    pre_content = "".join(b.get("content","") for b in blocks[:first_tool_pos] if b.get("type")=="content")
                    post_parts = [b.get("content","") for b in blocks[last_tool_pos+1:] if b.get("type")=="content"]
                    post_content = "".join(post_parts) if post_parts else (content or "")
                    tcs=[]
                    for idx, tc in enumerate(msg.get("tool_calls") or []):
                        args = tc.get("arguments") or {}
                        if not isinstance(args, dict):
                            args={}
                        tid = tc.get("id") or f"{tc.get('name','tool')}_{idx}"
                        if idx < len(tool_blocks) and tool_blocks[idx].get("id"):
                            tid = tool_blocks[idx]["id"]
                        tcs.append({"id": tid, "type":"function","function":{"name":tc.get("name","unknown"),"arguments":__import__("json").dumps(args)}})
                    llm_msg2 = {"role":"assistant","content":pre_content}
                    if msg.get("thinking"):
                        llm_msg2["reasoning_content"]=msg["thinking"]
                    if tcs:
                        llm_msg2["tool_calls"]=tcs
                    llm_messages.append(llm_msg2)
                    for idx, b in enumerate(tool_blocks):
                        tid = b.get("id") or (tcs[idx]["id"] if idx < len(tcs) else f"{b.get('name','tool')}_{idx}")
                        try:
                            cres = __import__("json").dumps(b.get("result"), default=str) if b.get("result") is not None else "No result"
                        except Exception:
                            cres = str(b.get("result") or "No result")
                        llm_messages.append({"role":"tool","content":cres,"tool_call_id":tid})
                        # Vision replay for context preview (so token counting is accurate)
                        try:
                            _pimgs = (b.get("result") or {}).get("images") if isinstance(b.get("result"), dict) else None
                            if _pimgs:
                                _pparts = [{"type":"text","text": f"[Tool {b.get('name','tool')} returned {len(_pimgs)} image(s)]"}]
                                for _pim in _pimgs:
                                    _pb64 = _pim.get("base64") or _pim.get("data") or ""
                                    _pmime = _pim.get("mime_type") or _pim.get("mimeType") or "image/png"
                                    if _pb64:
                                        _pparts.append({"type":"image_url","image_url":{"url": f"data:{_pmime};base64,{_pb64}"}})
                                if len(_pparts) > 1:
                                    llm_messages.append({"role":"user","content": _pparts})
                        except Exception:
                            pass
                    if post_content and post_content.strip():
                        llm_messages.append({"role":"assistant","content":post_content})
                llm_messages.insert(0, {"role": "system", "content": system_prompt_content})
            mcp_tools = []
            if mcp_manager:
                try:
                    mcp_tools = await mcp_manager.list_all_tools(include_disabled=False)
                except Exception:
                    mcp_tools = []
            enabled_mcp_servers = agent_config.get("enabled_mcp_servers") if agent_config else []
            if enabled_mcp_servers:
                allowed = set(enabled_mcp_servers)
                mcp_tools = [t for t in mcp_tools if t.get("server") in allowed]
            exclude_tools = []
            effective_enable_rag = False
            if agent_config and agent_config.get("enabled_tools"):
                enabled_custom = set(agent_config["enabled_tools"])
                if "query_documents" not in enabled_custom:
                    effective_enable_rag = False
                for ct in custom_tool_names():
                    if ct not in enabled_custom:
                        exclude_tools.append(ct)
            all_tools = tool_executor.get_tool_definitions(
                exclude_tools=exclude_tools,
                mcp_tools=mcp_tools,
                enable_rag=effective_enable_rag
            )
        except Exception as e:
            print(f"[CONTEXT] preview tools failed: {e}")
            llm_messages = []
            all_tools = []
        model_preview = (agent_config.get("model") if agent_config and agent_config.get("model") else "default")
        return {"context": {
            "model": model_preview,
            "system_prompt": system_prompt_content,
            "message_count": len(llm_messages),
            "tool_count": len(all_tools),
            "tools": [t.get("function", {}).get("name", "?") for t in all_tools],
        }}


@app.post("/api/conversations/{conversation_id}/messages")
async def send_message(conversation_id: str, request: Request):
    """Send a message and get LLM response"""
    data = await request.json()
    user_message = data.get("message", "")
    enable_rag = data.get("enable_rag", False)
    document_ids = data.get("document_ids")
    files = data.get("files", [])  # [{url, filename, type, size}]

    if not user_message.strip() and not files:
        raise HTTPException(status_code=400, detail="Message cannot be empty")

    # Steering support: if a stream is active for this conversation (the model
    # is mid-response), wait for it to finalize — the partial assistant message
    # is saved on cancel — BEFORE appending this user message, so DB order stays
    # correct: [user, assistant(partial), user(steering), assistant(new)].
    for _ in range(100):  # up to ~10s; the client aborts the stream first
        async with _streams_lock:
            active = bool(_active_streams.get(conversation_id))
        if not active:
            break
        await asyncio.sleep(0.1)

    # Build message content with file references included
    message_content = user_message
    
    async with get_db() as db:
        await add_message(db, conversation_id, "user", message_content, extra_metadata={"files": files} if files else None)
        request_id = str(uuid.uuid4())

        return {
            "request_id": request_id,
            "status": "processing",
            "enable_rag": enable_rag,
            "document_ids": document_ids,
            "files": files,
            "provider_id": data.get("provider_id")
        }
@app.post("/api/conversations/{conversation_id}/steer")
async def queue_steer(conversation_id: str, request: Request):
    """Queue a KV-preserving steer for an active stream.

    Frontend calls this *before* aborting SSE. The active _core_stream_handler
    sees the flag at the next chunk boundary, breaks gracefully (clean
    aclose on the llama.cpp stream -> slot KV retained), flushes the partial
    assistant message, and fast-unregisters. The caller then sends the real
    steering message via POST /messages as a normal follow-up whose prompt
    prefix reuses the cached context.
    """
    try:
        data = await request.json() if request.headers.get("content-type","").startswith("application/json") else {}
    except Exception:
        data = {}
    # Mark pending regardless of whether a stream is currently active — the
    # handler checks at chunk boundary; if no stream is active this is a no-op
    # and the normal /messages flow handles ordering.
    await _set_steer_pending(conversation_id)
    # Also ensure the active-stream poll in /messages will unblock quickly:
    # we do NOT append the steering message here; frontend will POST it
    # separately after the grace window so ordering stays DB-correct.
    return {"queued": True, "conversation_id": conversation_id}






async def _generate_title_with_model(
    llm_messages: list,
    assistant_message: str,
    llm_client,
    tools: list = None,
    model: str = None,
    base_url: str = None,
    api_key: str = None,
    thinking_content: str = None,
    thinking_mode: str = None,
    message_blocks: list = None,
) -> str:
    """
    Generate a title by appending to the existing conversation messages.
    KV cache reuses the prefix from the main response.
    The appended messages are not saved to DB — only the title is returned.
    Phase A: uses expanded final (tool-aware) and same thinking_mode as main
    so the next turn's prompt shares the full prefix (98% hit).
    """
    title_messages = list(llm_messages)
    # Phase A: expand final like history replay, not single
    if message_blocks is not None:
        title_messages.extend(_expand_final_for_autos(assistant_message, thinking_content, message_blocks))
    else:
        title_assistant = {"role": "assistant", "content": assistant_message}
        if thinking_content:
            title_assistant["reasoning_content"] = thinking_content
        title_messages.append(title_assistant)
    title_messages.append({"role": "user", "content": "Generate a title (3-6 words) for this conversation. No reasoning. Just output the title."})

    title = ""
    thinking = ""
    try:
        async for chunk in llm_client.stream_chat(
            title_messages,
            model=model,
            temperature=0.0,
            max_tokens=1024,
            tools=tools,
            tool_choice=None,
            base_url=base_url,
            api_key=api_key,
            thinking_mode=thinking_mode,
        ):
            chunk_type = chunk.get("type")
            if chunk_type == "content":
                title += chunk.get("content", "")
            elif chunk_type == "thinking":
                thinking += chunk.get("content", "")
            else:
                print(f"[TITLE GEN] unexpected chunk type: {chunk_type}")
    except Exception as e:
        print(f"Title generation error: {e}")
        return ""

    print(f"[TITLE GEN] raw title: '{title[:200]}'")
    print(f"[TITLE GEN] thinking length: {len(thinking)} chars")

    # If content is empty but thinking has content, the title might be inside thinking
    if not title.strip() and thinking.strip():
        title = thinking

    import re
    title = re.sub(r'<think>.*?</think>', '', title, flags=re.DOTALL)
    title = re.sub(r'<[^>]+>', '', title).strip()
    words = title.split()
    if not words:
        return ""
    title = ' '.join(words[:6]).strip().rstrip('.,;:!?\\-"\'')
    return title[:60]


async def _maybe_reflect_and_propose_skill(db, conversation_id: str, llm_client, model=None,
                                           base_url=None, api_key=None,
                                           llm_messages=None, assistant_message=None, thinking_content=None,
                                           tools=None, thinking_mode: str = None, message_blocks: list = None):
    """Self-improvement loop (Phase 4): propose a skill draft after a
    multi-tool task.

    Insight-based: runs whenever the last turn used tools; the LLM decides
    if the procedure is worth turning into a skill (no fixed cadence).
    Prompt reuses the conversation prefix (llm_messages) for KV cache hit
    when caller provides it — otherwise falls back to a standalone prompt.
    Output is always a DRAFT under skills/_drafts/.

    Returns a dict describing what happened for auto_action visibility:
    {"action":"skill","status":"completed|skipped|error","detail":{...}}
    """
    try:
        from settings import settings_manager as _sm2
        _iv = _sm2.get_settings().get("memory_auto_extract_interval", 1)
        if isinstance(_iv, int) and _iv <= 0:
            return {"action": "skill", "status": "skipped", "detail": {"reason": "skill reflection disabled"}}
        # Only reflect when tools were used in the last assistant message.
        msgs = await get_conversation_messages(db, conversation_id)
        last_msg = msgs[-1] if msgs else None
        blocks = (last_msg.get("metadata") or {}).get("blocks") or [] if last_msg else []
        tool_names = [b.get("name") for b in blocks if b.get("type") == "tool_call"]
        if not tool_names:
            return {"action": "skill", "status": "skipped", "detail": {"reason": "no tool calls in last turn"}}
        last_user = next((m["content"] for m in reversed(msgs) if m["role"] == "user"), None)
        last_assistant = next((m["content"] for m in reversed(msgs) if m["role"] == "assistant"), None)
        if not last_user or not last_assistant:
            return {"action": "skill", "status": "skipped", "detail": {"reason": "no exchange found"}}
        prompt = (
            "You are a skill-builder for an AI assistant. Decide whether this task is worth "
            "turning into a reusable skill.\n\n"
            f"USER: {last_user[:1500]}\n\nASSISTANT: {last_assistant[:1500]}\n\n"
            f"TOOLS USED: {', '.join(tool_names[:10])}\n\n"
            "Rules:\n"
            "- CREATE a skill when the task was a repeatable, multi-step procedure "
            "(at least 2 concrete steps).\n"
            "- IMPROVE an existing skill only when load_skill was used and the run "
            "seems to have gone wrong or could clearly be better.\n"
            "- Otherwise output none.\n"
            'Output ONLY a JSON object: {"action": "create"|"improve"|"none", '
            '"name": "short-slug", "description": "one line", '
            '"instructions": "numbered steps", "reason": "short justification"}'
        )
        # KV-friendly: reuse prefix (like title generation) when caller provides it
        # Phase A: use expanded final (tool-aware) and same thinking_mode/tool_choice as main
        if llm_messages is not None and assistant_message is not None:
            _msgs = list(llm_messages)
            if message_blocks is not None:
                _msgs.extend(_expand_final_for_autos(assistant_message, thinking_content, message_blocks))
            else:
                _assistant = {"role": "assistant", "content": assistant_message}
                if thinking_content:
                    _assistant["reasoning_content"] = thinking_content
                _msgs.append(_assistant)
            _msgs.append({"role": "user", "content": prompt})
            stream_target = _msgs
            _tools = tools
            _tool_choice = None  # match main (None, not "none") so prompt tokens identical
        else:
            stream_target = [{"role": "user", "content": prompt}]
            _tools = None
            _tool_choice = None
        raw = ""
        async for chunk in llm_client.stream_chat(
            stream_target,
            model=model, temperature=0.0, max_tokens=700, tools=_tools,
            tool_choice=_tool_choice,
            base_url=base_url, api_key=api_key,
            thinking_mode=thinking_mode,
        ):
            if chunk.get("type") == "content":
                raw += chunk.get("content", "")
        import re as _re
        import json as _json
        m = _re.search(r"\{.*\}", raw, _re.DOTALL)
        if not m:
            print("[SKILLS] reflection: no JSON in model output")
            return {"action": "skill", "status": "skipped", "detail": {"reason": "no JSON in model output", "raw": raw[:500]}}
        try:
            decision = _json.loads(m.group(0))
        except _json.JSONDecodeError as e:
            print(f"[SKILLS] reflection: unparseable JSON: {e}")
            return {"action": "skill", "status": "error", "detail": {"reason": f"unparseable JSON: {e}", "raw": raw[:500]}}
        action = str(decision.get("action") or "none").strip().lower()
        if action not in ("create", "improve"):
            print(f"[SKILLS] reflection: action={action} — no draft")
            return {"action": "skill", "status": "skipped", "detail": {"reason": f"model chose none ({action})", "decision": decision}}
        name = str(decision.get("name") or "").strip()
        description = str(decision.get("description") or "").strip()
        instructions = str(decision.get("instructions") or "").strip()
        reason = str(decision.get("reason") or "").strip()
        if not name or not instructions:
            print("[SKILLS] reflection: missing name/instructions")
            return {"action": "skill", "status": "error", "detail": {"reason": "missing name/instructions", "decision": decision}}
        from tools.skills_tool import write_skill, get_skill
        if action == "improve" and not get_skill(name):
            print(f"[SKILLS] reflection: improve target '{name}' does not exist — skipping")
            return {"action": "skill", "status": "skipped", "detail": {"reason": f"improve target '{name}' not found"}}
        body = instructions
        if reason:
            body = f"<!-- reflection reason: {reason} -->\n\n" + body
        skill = write_skill(name, description or name, body, draft=True)
        print(f"[SKILLS] reflection: draft proposed: {name} ({action}) — review in Skills modal")
        return {"action": "skill", "status": "completed", "detail": {"skill": name, "skill_action": action, "description": description, "reason": reason}}
    except Exception as e:
        print(f"[SKILLS] reflection failed: {e}")
        import traceback as _tb
        _tb.print_exc()
        return {"action": "skill", "status": "error", "detail": {"reason": str(e)[:500]}}


async def _extract_memory_from_exchange(db, conversation_id: str, llm_client, agent_id=None, model=None,
                                        base_url=None, api_key=None,
                                        llm_messages=None, assistant_message=None, thinking_content=None,
                                        tools=None, thinking_mode: str = None, message_blocks: list = None):
    """Auto-extract durable facts from the last user↔assistant exchange (Phase 2).

    Runs every N user turns (settings.memory_auto_extract_interval; <=0 disables).
    The LLM decides whether the exchange holds durable insights; each candidate
    fact is then checked against the existing store (FTS + token-Jaccard dedup)
    so regenerations and follow-up turns cannot re-add the same fact.

    When `llm_messages` is provided, the prompt is appended to the existing
    prefix for KV cache reuse (like title generation).
    """
    try:
        from settings import settings_manager as _sm
        # memory_auto_extract_interval = run every N user turns (<=0 disables).
        # Cadence bounds extraction cost and duplicate pressure; the dedup guard
        # below covers the residual overlap.
        _interval = _sm.get_settings().get("memory_auto_extract_interval", 3)
        if not isinstance(_interval, int) or isinstance(_interval, bool):
            _interval = 3
        if _interval <= 0:
            return {"action": "memory", "status": "skipped", "detail": {"reason": "memory extraction disabled"}}
        msgs = await get_conversation_messages(db, conversation_id)
        if _interval > 1:
            _user_turns = sum(1 for m in msgs if m["role"] == "user")
            if _user_turns % _interval != 0:
                return {"action": "memory", "status": "skipped", "detail": {"reason": f"cadence: runs every {_interval} user turns (turn {_user_turns})"}}
        last_user = next((m["content"] for m in reversed(msgs) if m["role"] == "user"), None)
        last_assistant = next((m["content"] for m in reversed(msgs) if m["role"] == "assistant"), None)
        if not last_user or not last_assistant:
            return {"action": "memory", "status": "skipped", "detail": {"reason": "no exchange found"}}
        # Show the extractor what is already saved so it can skip covered ground
        # (write-side dedup below is the safety net, not the primary mechanism).
        from database.memory_crud import fts_search_memory
        try:
            _existing = await fts_search_memory(db, last_user, top_k=6)
        except Exception:
            _existing = []
        _existing_block = "\n".join(f"- {e['content']}" for e in _existing) or "(none)"
        prompt = (
            "Extract durable insights worth remembering from this conversation exchange.\n"
            "ALREADY SAVED (skip anything these already cover):\n"
            f"{_existing_block}\n\n"
            "Rules:\n"
            "- Save ONLY facts that stay useful months from now: user preferences and corrections, identity/context, project decisions, environment/tool quirks, durable workflow lessons.\n"
            "- Do NOT save in-progress task state ('pending', 'awaiting review'), one-off actions or logs ('cleanup done', 'deleted N files'), or anything only this transcript shows.\n"
            "- Never store secret VALUES (API keys, passwords, tokens) — at most where they are kept.\n"
            "- Each fact: one complete standalone sentence (15-160 chars, contains a verb) a stranger could understand. No single words or tags like 'llm-ui', 'skills', 'tools', 'architecture'.\n"
            "- Give each fact 1-3 short kebab-case tags (e.g. \"finance\", \"verit-analytics\", \"tts\").\n"
            'Output ONLY a JSON array of objects: {"fact": "...", "tags": ["..."]}. If nothing durable, output [].\n\n'
            f"User: {last_user[:1500]}\n\nAssistant: {last_assistant[:1500]}"
        )
        # KV-friendly path: append to existing prefix (like title generation)
        # Phase A: use expanded final (tool-aware) and same thinking_mode/tool_choice as main
        if llm_messages is not None and assistant_message is not None:
            _msgs = list(llm_messages)
            if message_blocks is not None:
                _msgs.extend(_expand_final_for_autos(assistant_message, thinking_content, message_blocks))
            else:
                _assistant = {"role": "assistant", "content": assistant_message}
                if thinking_content:
                    _assistant["reasoning_content"] = thinking_content
                _msgs.append(_assistant)
            _msgs.append({"role": "user", "content": prompt})
            stream_target = _msgs
            _tools = tools
            _tool_choice = None  # match main (None, not "none")
        else:
            stream_target = [{"role": "user", "content": prompt}]
            _tools = None
            _tool_choice = None
        parts = []
        async for chunk in llm_client.stream_chat(
            stream_target,
            model=model, temperature=0.0, max_tokens=512, tools=_tools,
            tool_choice=_tool_choice,
            base_url=base_url, api_key=api_key,
            thinking_mode=thinking_mode,
        ):
            if chunk.get("type") == "content":
                parts.append(chunk.get("content", ""))
        raw = "".join(parts)
        import re
        import json as _json

        def _sanitize(s: str) -> str:
            # Fallback line-parsing used to store raw JSON lines verbatim:
            # `"some fact…",` — strip quotes and trailing commas.
            s = s.strip().rstrip(",").strip()
            while len(s) >= 2 and s[0] == '"' and s[-1] == '"':
                s = s[1:-1].strip().rstrip(",").strip()
            return s

        # --- Parse: JSON objects {"fact","tags"} preferred, bare strings tolerated ---
        items = []  # list of (fact, tags)
        m = re.search(r"\[.*\]", raw, re.DOTALL)
        parsed = None
        if m:
            try:
                parsed = _json.loads(m.group(0))
            except _json.JSONDecodeError:
                parsed = None
        if isinstance(parsed, list):
            for x in parsed:
                if isinstance(x, dict):
                    fact = _sanitize(str(x.get("fact") or x.get("content") or ""))
                    tags = [str(t).strip().lower() for t in (x.get("tags") or []) if str(t).strip()][:3]
                else:
                    fact, tags = _sanitize(str(x or "")), []
                if fact:
                    items.append((fact, tags))
        else:
            for l in raw.splitlines():
                t = l.strip()
                if not t or t.startswith("```"):
                    continue
                fact = _sanitize(t.lstrip("- ").strip('"'))
                if fact and len(fact) > 5 and not fact.startswith("{"):
                    items.append((fact, []))
        # --- Filter single-word/tag junk (FTS would index these, but useless) ---
        _filtered = []
        for _f, _tags in items:
            if len(_f) < 15 or _f.count(" ") < 2:
                print(f"[MEMORY] filtered short/single-word: '{_f[:60]}'")
                continue
            if _f.lower() in {"llm-ui", "llm_ui", "skills", "tools", "architecture"}:
                print(f"[MEMORY] filtered known tag: '{_f}'")
                continue
            _filtered.append((_f, _tags))
        items = _filtered
        if not items:
            return {"action": "memory", "status": "skipped", "detail": {"reason": "no durable facts found (filtered)"}}
        from database.memory_crud import create_memory_entry_dedup
        scope = f"agent:{agent_id}" if agent_id is not None else "global"
        added, skipped = [], 0
        for fact, tags in items[:10]:
            entry, dup = await create_memory_entry_dedup(db, fact, scope=scope, tags=tags, source="auto")
            if dup:
                skipped += 1
                print(f"[MEMORY] dedup skip (≈ {dup['id'][:8]}): {fact[:70]}")
            else:
                added.append(fact)
        if added:
            await db.commit()
            print(f"[MEMORY] auto-extracted {len(added)} fact(s), {skipped} dedup-skipped (scope={scope})")
            return {"action": "memory", "status": "completed",
                    "detail": {"facts": added, "count": len(added), "skipped_duplicates": skipped, "scope": scope}}
        return {"action": "memory", "status": "skipped", "detail": {"reason": f"all {skipped} candidate(s) already in memory"}}
    except Exception as e:
        print(f"[MEMORY] extraction failed: {e}")
        import traceback as _tb2
        _tb2.print_exc()
        return {"action": "memory", "status": "error", "detail": {"reason": str(e)[:500]}}


async def _stream_with_stall_timeout(generator, initial_timeout=600, stall_timeout=60):
    """Wrap an async generator with two-phase stall detection.

    Replaces a total-time watchdog so the LLM can generate for arbitrarily long
    responses as long as it keeps producing chunks. Compatible with both MTP
    and non-MTP models.

    Two phases:
      1. Initial: generous timeout (default 600s) to allow for long prompt
         processing / prefill before the first chunk arrives.
      2. Stall: shorter timeout (default 60s) for subsequent chunks — catches
         genuine stalls where the model stops generating.

    Args:
        generator: An async generator (e.g. llm_client.stream_chat(...)).
        initial_timeout: Max seconds to wait for the FIRST chunk (prompt processing).
        stall_timeout: Max seconds to wait for the NEXT chunk after the first one.

    Yields:
        Each chunk from the wrapped generator.
    """
    first_chunk = True
    try:
        while True:
            timeout = initial_timeout if first_chunk else stall_timeout
            chunk = await asyncio.wait_for(
                generator.__anext__(),
                timeout=timeout
            )
            first_chunk = False
            yield chunk
    except StopAsyncIteration:
        pass
    except asyncio.TimeoutError:
        which = "initial (prompt processing)" if first_chunk else "inter-chunk"
        print(f"[WATCHDOG] Stream stalled for {timeout}s ({which}) — no chunks received")


# ── Steering: KV-preserving graceful interrupt ─────────────────────────
# Frontend queues a steer (POST /api/conversations/{id}/steer) instead of
# hard-aborting SSE. The active _core_stream_handler sees the flag at the
# next chunk boundary, breaks gracefully (clean aclose -> llama.cpp slot KV
# retained), flushes the partial, and unregisters fast. The steer message
# is then sent as a normal follow-up whose prompt reuses the full prefix.
_steer_pending: Dict[str, float] = {}
_steer_lock = asyncio.Lock()


async def _set_steer_pending(conversation_id: str) -> None:
    async with _steer_lock:
        _steer_pending[conversation_id] = _time.monotonic()
        print(f"[STEER] queued for {conversation_id[:8]}")


async def _consume_steer_pending(conversation_id: str) -> bool:
    async with _steer_lock:
        if conversation_id in _steer_pending:
            _steer_pending.pop(conversation_id, None)
            return True
        return False


async def _is_steer_pending(conversation_id: str) -> bool:
    async with _steer_lock:
        return conversation_id in _steer_pending


async def _clear_steer_pending(conversation_id: str) -> None:
    async with _steer_lock:
        _steer_pending.pop(conversation_id, None)


# ── Active stream registry (steering support) ──────────────────────────────
# Tracks which request_ids are streaming per conversation so a new user
# message (steering) can wait for the active response to finalize first —
# guaranteeing the partial assistant message is saved BEFORE the steering
# message in the DB, keeping message order correct.
_active_streams: Dict[str, set] = {}
_streams_lock = asyncio.Lock()


async def _register_stream(conversation_id: str, request_id: str) -> None:
    try:
        async with _streams_lock:
            _active_streams.setdefault(conversation_id, set()).add(request_id)
    except Exception:
        pass


async def _unregister_stream(conversation_id: str, request_id: str) -> None:
    try:
        async with _streams_lock:
            s = _active_streams.get(conversation_id)
            if s:
                s.discard(request_id)
                if not s:
                    del _active_streams[conversation_id]
    except Exception:
        pass


async def _save_assistant_message(db, conversation_id: str, assistant_message: str,
                                  thinking_content: str, message_blocks: list,
                                  model: Optional[str], version: int,
                                  version_group: Optional[str],
                                  turn_index: Optional[float] = None,
                                  metrics: Optional[dict] = None):
    """Persist an assistant message. Returns saved message dict or None."""
    if not (assistant_message.strip() or thinking_content.strip() or message_blocks):
        return None
    # Consolidate consecutive content/thinking blocks (preserve exact formatting)
    consolidated_blocks = []
    for block in message_blocks:
        btype = block.get('type')
        if btype in ('content', 'thinking'):
            if consolidated_blocks and consolidated_blocks[-1].get('type') == btype:
                consolidated_blocks[-1]['content'] = (
                    consolidated_blocks[-1].get('content', '') + block.get('content', '')
                )
            else:
                consolidated_blocks.append(block)
        else:
            consolidated_blocks.append(block)
    message_extra_metadata = {"model": model} if model else {}
    if metrics:
        message_extra_metadata["metrics"] = metrics
    saved = await add_message(
        db, conversation_id, "assistant", assistant_message,
        blocks=consolidated_blocks or None,
        extra_metadata=message_extra_metadata,
        version=version,
        version_group=version_group,
        turn_index=turn_index
    )
    await db.commit()
    return saved



def _aggregate_turn_metrics(turn_metrics: List[Dict]) -> Optional[dict]:
    """Collapse per-LLM-call metrics for one turn into what we persist.

    Single call: its own metrics. Tool-loop turns: sum counts/durations, keep the
    first TTFT, and retain `_iterations` so the UI can recover the real final
    context occupancy (a summed prompt_tokens double-counts the re-sent prefix).
    Returns None for a turn that never completed an LLM call.
    """
    if not turn_metrics:
        return None
    if len(turn_metrics) == 1:
        return turn_metrics[0]
    prompt_sum = sum((m.get("prompt_tokens") or 0) for m in turn_metrics)
    compl_sum = sum((m.get("completion_tokens") or 0) for m in turn_metrics)
    total_sum = sum((m.get("total_tokens") or 0) for m in turn_metrics)
    ttft_first = next((m.get("ttft_ms") for m in turn_metrics if m.get("ttft_ms") is not None), None)
    dur_sum = sum((m.get("total_duration_ms") or 0) for m in turn_metrics)
    cached_sum = None
    if any(m.get("cached_tokens") is not None for m in turn_metrics):
        cached_sum = sum((m.get("cached_tokens") or 0) for m in turn_metrics)
    agg_tps = None
    gen_ms = (dur_sum - (ttft_first or 0)) if ttft_first else dur_sum
    if compl_sum and gen_ms > 0:
        agg_tps = round(compl_sum / (gen_ms / 1000), 2)
    final = {
        "ttft_ms": ttft_first,
        "total_duration_ms": dur_sum,
        "prompt_tokens": prompt_sum or None,
        "completion_tokens": compl_sum or None,
        "total_tokens": total_sum or (prompt_sum + compl_sum if prompt_sum or compl_sum else None),
        "cached_tokens": cached_sum,
        "tokens_per_second": agg_tps,
        "prompt_per_second": turn_metrics[0].get("prompt_per_second"),
    }
    final["_iterations"] = turn_metrics
    return final


def _expand_final_for_autos(
    assistant_message: str, thinking_content: str, message_blocks: List[Dict]
) -> List[Dict]:
    """
    Phase A KV fix: autos must append the *same* token sequence as the
    next turn's history will replay. Main history replays a tool turn as
    `assistant(tool_calls,pre_content) -> tool(s) -> assistant(post_content)`
    (see `llm_messages` build). Previously autos appended a single
    `assistant(assistant_message)` → prefix mismatch 80% vs 98%.
    This helper returns 1 or 3 messages matching the replay.
    """
    if not message_blocks:
        m: Dict = {"role": "assistant", "content": assistant_message or ""}
        if thinking_content:
            m["reasoning_content"] = thinking_content
        return [m]
    # Find tool blocks in the current turn's message_blocks
    tool_blocks = [b for b in message_blocks if b.get("type") == "tool_call"]
    if not tool_blocks:
        m2: Dict = {"role": "assistant", "content": assistant_message or ""}
        if thinking_content:
            m2["reasoning_content"] = thinking_content
        return [m2]
    # Split like history replay: pre-content/thinking before first tool, post after last
    first_pos = next((i for i, b in enumerate(message_blocks) if b.get("type") == "tool_call"), len(message_blocks))
    last_pos = len(message_blocks) - 1 - next((i for i, b in enumerate(reversed(message_blocks)) if b.get("type") == "tool_call"), len(message_blocks))
    pre_content = "".join(b.get("content", "") for b in message_blocks[:first_pos] if b.get("type") == "content")
    pre_thinking = "".join(b.get("content", "") for b in message_blocks[:first_pos] if b.get("type") == "thinking")
    post_parts = [b.get("content", "") for b in message_blocks[last_pos + 1 :] if b.get("type") == "content"]
    post_content = "".join(post_parts) if post_parts else (assistant_message or "")
    post_thinking_parts = [b.get("content", "") for b in message_blocks[last_pos + 1 :] if b.get("type") == "thinking"]
    post_thinking = "".join(post_thinking_parts)
    if not post_thinking and thinking_content and not pre_thinking:
        # Fallback: thinking_content belongs to post when no pre-thinking found
        post_thinking = thinking_content
        pre_thinking = pre_thinking or ""
    elif thinking_content and not pre_thinking and not post_thinking:
        pre_thinking = thinking_content

    # Build tcs from tool_blocks (preserve stored ids)
    tcs: List[Dict] = []
    for idx, b in enumerate(tool_blocks):
        tid = b.get("id") or f"{b.get('name','tool')}_{idx}"
        args = b.get("arguments") or {}
        if not isinstance(args, dict):
            args = {}
        tcs.append({
            "id": tid,
            "type": "function",
            "function": {"name": b.get("name", "unknown"), "arguments": __import__("json").dumps(args)},
        })

    out: List[Dict] = []
    first: Dict = {"role": "assistant", "content": pre_content}
    if pre_thinking:
        first["reasoning_content"] = pre_thinking
    if tcs:
        first["tool_calls"] = tcs
    out.append(first)
    for idx, b in enumerate(tool_blocks):
        tid = b.get("id") or (tcs[idx]["id"] if idx < len(tcs) else f"{b.get('name','tool')}_{idx}")
        result = b.get("result")
        try:
            cres = __import__("json").dumps(result, default=str) if result is not None else "No result"
        except Exception:
            cres = str(result) if result is not None else "No result"
        out.append({"role": "tool", "content": cres, "tool_call_id": tid})
    if (post_content and post_content.strip()) or post_thinking:
        last: Dict = {"role": "assistant", "content": post_content}
        if post_thinking:
            last["reasoning_content"] = post_thinking
        out.append(last)
    return out




async def _persist_partial_turn(conversation_id: str, assistant_message: str,
                                thinking_content: str, message_blocks: List[Dict],
                                model: Optional[str], version: int,
                                version_group: Optional[str],
                                turn_index: Optional[float],
                                placeholder_id: Optional[str],
                                turn_metrics: Optional[List[Dict]] = None) -> Optional[str]:
    """Best-effort persist of a partial turn on crash/cancel/error, using a
    fresh DB session (the generator's session is gone by then). Returns the
    saved row id when known so clients can repair their local placeholder."""
    try:
        _fm = _aggregate_turn_metrics(turn_metrics or [])
        async with get_db() as pdb:
            if placeholder_id:
                cons = []
                for b in message_blocks:
                    bt = b.get("type")
                    if bt in ("content", "thinking"):
                        if cons and cons[-1].get("type") == bt:
                            cons[-1]["content"] = cons[-1].get("content", "") + b.get("content", "")
                        else:
                            cons.append(dict(b))
                    else:
                        cons.append(dict(b))
                full_c = "".join(x.get("content", "") for x in cons if x.get("type") == "content") or assistant_message
                full_t = "".join(x.get("content", "") for x in cons if x.get("type") == "thinking") or thinking_content
                _extra: Dict = {}
                if model:
                    _extra["model"] = model
                if _fm:
                    _extra["metrics"] = _fm
                if full_c.strip() or full_t.strip() or cons:
                    await update_assistant_message_full(
                        pdb, placeholder_id,
                        content=full_c,
                        thinking=full_t,
                        blocks=cons or message_blocks,
                        extra_metadata=_extra or None,
                    )
                return placeholder_id
            saved = await _save_assistant_message(
                pdb, conversation_id, assistant_message, thinking_content,
                message_blocks, model, version, version_group,
                turn_index=turn_index,
                metrics=_fm
            )
            return saved["id"] if isinstance(saved, dict) else None
    except Exception as e:
        print(f"[STREAM] partial save on teardown failed: {e}")
        import traceback as _tb_p
        _tb_p.print_exc()
        return None
async def _core_stream_handler(
    request_id: str,
    conversation_id: str,
    enable_rag: bool = False,
    model: Optional[str] = None,
    document_ids: Optional[list] = None,
    version: int = 1,
    version_group: Optional[str] = None,
    anchor_message_id: Optional[str] = None,
    turn_index: Optional[float] = None,
    provider_id: Optional[str] = None,
    override_servers: Optional[list] = None,
    thinking_mode: Optional[str] = None
) -> AsyncGenerator[str, None]:
    """Universal SSE handler for streaming LLM responses and tool execution.

    Args:
        version: Version number for regenerated responses
        version_group: UUID shared by all versions of the same response
        anchor_message_id: Regeneration only — the user message being re-answered.
                           History replay is cut at (and includes) this message and
                           every stored version of the response group is excluded, so
                           the model never sees prior attempts or later turns.
        turn_index: Regeneration only — timeline slot inherited from the superseded
                    response so all versions render in place.
        provider_id: LLM provider to use (falls back to the default provider)
    """
    current_document_ids.set(document_ids)

    # Placeholder for crash-safe incremental persist — must be defined here
    # so the CancelledError handler (outside the `async with get_db()` block)
    # can still access the current turn's state even if cancelled early.
    placeholder_id: Optional[str] = None
    assistant_message = ""
    thinking_content = ""
    message_blocks: List[Dict] = []
    # Per-LLM-call metrics for this turn. Declared here (not inside the DB block)
    # so the cancel/error teardown handlers always see whatever completed calls
    # recorded — those turns used to be saved with no metrics at all.
    turn_metrics: List[Dict] = []
    try:
        await _register_stream(conversation_id, request_id)
        async with get_db() as db:
            # Get conversation to retrieve agent configuration
            from sqlalchemy import select
            from database.models import Conversation
            result = await db.execute(select(Conversation).where(Conversation.id == conversation_id))
            conversation = result.scalar_one_or_none()

            # Get agent configuration if conversation has an agent
            agent_config = None
            if conversation:
                agent_id = conversation.agent_id
                if agent_id is not None:
                    agent = await get_agent(db, agent_id)
                    if agent:
                        agent_config = {
                            "system_prompt": agent.system_prompt,
                            "model": agent.model,
                            "provider_id": getattr(agent, "provider_id", None),
                            "temperature": agent.temperature,
                            "top_k": agent.top_k,
                            "max_tokens": agent.max_tokens,
                            "enable_rag": bool(agent.enable_rag),
                            "enabled_tools": agent.enabled_tools or [],
                            "enabled_mcp_servers": agent.enabled_mcp_servers or [],
                            "enabled_skills": agent.enabled_skills or []
                        }

            # Agent sampling reaches the LLM here. Values are taken verbatim from
            # the agent row; 0 / None means "not configured" so the client keeps
            # falling back to the settings defaults (previous behaviour).
            _ac = agent_config or {}
            _agent_temperature = _ac.get("temperature") or None
            _agent_top_k = _ac.get("top_k") or None
            _agent_max_tokens = _ac.get("max_tokens") or None

            # Resolve LLM provider — single entity: model implies provider (live chat overrides default/agent)
            # Default from settings is fallback only when no live model selected
            from database.provider_crud import get_provider as _get_provider, get_default_provider, list_providers as _list_providers
            provider = None
            provider_base_url = None
            provider_api_key = None
            # 1) If a live model is selected (chat dropdown), it is single entity → find its owner provider
            if model:
                try:
                    _all = await _list_providers(db, include_api_key=True)
                    for p in _all:
                        if not p.get("enabled"):
                            continue
                        if any(m.get("id") == model for m in (p.get("models") or [])):
                            provider = p
                            provider_base_url = p.get("base_url")
                            provider_api_key = p.get("api_key")
                            print(f"[PROVIDER] live model '{model}' → provider '{p['name']}' ({provider_base_url}) [single entity]")
                            break
                    if provider is None:
                        print(f"[PROVIDER] live model '{model}' not in any enabled provider cache — will try agent/requested/default")
                except Exception as e:
                    print(f"[PROVIDER] live-model lookup failed: {e}")
            # 2) Fallback: agent's provider > requested provider_id > default (for default model from settings)
            if provider is None:
                resolved_provider_id = None
                if agent_config and agent_config.get("provider_id"):
                    resolved_provider_id = agent_config["provider_id"]
                elif provider_id:
                    resolved_provider_id = provider_id
                try:
                    if resolved_provider_id:
                        provider = await _get_provider(db, resolved_provider_id, include_api_key=True)
                    if provider is None:
                        provider = await get_default_provider(db, include_api_key=True)
                except Exception as e:
                    print(f"[PROVIDER] resolution failed: {e}")
                provider_base_url = provider.get("base_url") if provider else None
                provider_api_key = provider.get("api_key") if provider else None
                if provider:
                    print(f"[PROVIDER] using '{provider['name']}' ({provider_base_url}) model={model} [fallback]")
            # If no live model was supplied (race: page reloaded before model
            # list fetched, so selectedModel is empty), don't fall back to
            # the stale settings default (e.g. Qwen no longer deployed) — derive
            # from the resolved (or any enabled) provider instead.
            if not model and provider:
                cand = [m.get("id") for m in (provider.get("models") or []) if m.get("id")]
                if cand:
                    model = cand[0]
                    print(f"[PROVIDER] derived model '{model}' from provider '{provider['name']}' (live selection was empty)")
            if not model:
                try:
                    _all2 = await _list_providers(db)
                    for p in _all2:
                        if not p.get("enabled"):
                            continue
                        ms = [m.get("id") for m in (p.get("models") or []) if m.get("id")]
                        if ms:
                            model = ms[0]
                            provider = p
                            provider_base_url = p.get("base_url")
                            provider_api_key = p.get("api_key")
                            print(f"[PROVIDER] global fallback model '{model}'")
                            break
                except Exception:
                    pass
            # 3) Safety: if resolved provider still doesn't own live model (stale cache or agent pin), scan again
            if model and provider:
                prov_models = [m.get("id") for m in (provider.get("models") or []) if m.get("id")]
                if prov_models and model not in prov_models:
                    orig_name = provider.get("name")
                    switched = False
                    # honour explicit provider_id if it owns model
                    if provider_id and provider_id != provider.get("id"):
                        try:
                            req_p = await _get_provider(db, provider_id, include_api_key=True)
                            if req_p and any(m.get("id") == model for m in (req_p.get("models") or [])):
                                provider = req_p
                                provider_base_url = provider.get("base_url")
                                provider_api_key = provider.get("api_key")
                                print(f"[PROVIDER] honouring requested provider '{provider['name']}' for model '{model}' (was '{orig_name}')")
                                switched = True
                        except Exception as e:
                            print(f"[PROVIDER] requested-provider check failed: {e}")
                    if not switched:
                        try:
                            all_ps = await _list_providers(db, include_api_key=True)
                            for p in all_ps:
                                if not p.get("enabled"):
                                    continue
                                if any(m.get("id") == model for m in (p.get("models") or [])):
                                    if p.get("id") != provider.get("id"):
                                        print(f"[PROVIDER] switching from '{orig_name}' to '{p['name']}' for model '{model}' [stale cache]")
                                        provider = p
                                        provider_base_url = p.get("base_url")
                                        provider_api_key = p.get("api_key")
                                    switched = True
                                    break
                            if not switched:
                                print(f"[PROVIDER] model '{model}' not in any enabled provider cache — leaving '{orig_name}' (will 400 if missing)")
                        except Exception as e:
                            print(f"[PROVIDER] model-aware fallback failed: {e}")
            if provider:
                print(f"[PROVIDER] final '{provider['name']}' ({provider_base_url}) model={model} provider_id={provider.get('id')}")
            
            # Get current date for system prompt
            from datetime import datetime
            current_date = datetime.now().strftime("%Y-%m-%d")
            
            # Build system prompt with current date + app identity (local llm_ui)
            identity_hint = "You are running inside llm_ui — a local, single-user chat app on this machine. Conversations, memory, agents, skills and files you manage are all inside llm_ui unless the user explicitly names an external app. Memory: relevant durable facts may be appended to the user's message in a <relevant_memories> block — use them. Before answering about past work, preferences, projects or people, call memory_search. When the user states a durable fact or preference (or says 'remember X'), call memory_write immediately: concise, standalone, tagged. Never store secret values — at most where they are kept."
            system_prompt_content = agent_config["system_prompt"] if agent_config and agent_config.get("system_prompt") else ""
            if system_prompt_content:
                system_prompt_content = f"{system_prompt_content}\n\n{identity_hint}\n\nCurrent date: {current_date}"
            else:
                system_prompt_content = f"You are a helpful AI assistant. {identity_hint} Current date: {current_date}"

            # Inject persistent agent memory (Phase 2)
            try:
                from database.memory_crud import get_memory_for_injection
                memory_block = await get_memory_for_injection(
                    db,
                    agent_id=conversation.agent_id if conversation else None,
                    conversation_id=conversation_id
                )
                if memory_block:
                    system_prompt_content += "\n\n" + memory_block
            except Exception as e:
                print(f"[MEMORY] injection failed: {e}")

            # Inject skills index (Phase 3) — compact, one line per skill.
            try:
                from tools.skills_tool import skill_index
                idx = skill_index()
                if idx:
                    # Agent can restrict which skills are visible; empty = all.
                    enabled_skills = set()
                    if agent_config and agent_config.get("enabled_skills"):
                        enabled_skills = set(agent_config["enabled_skills"])
                    if enabled_skills:
                        idx = "\n".join(
                            line for line in idx.splitlines()
                            if line.startswith("- ")
                            and line[2:].split(":", 1)[0].strip() in enabled_skills
                        )
                    if idx:
                        system_prompt_content += (
                            "\n\n### Available skills\n"
                            "Call load_skill(name) to load a skill's full instructions "
                            "when it is relevant to the current task.\n" + idx
                        )
            except Exception as e:
                print(f"[SKILLS] index injection failed: {e}")
            
            messages = await get_conversation_messages(db, conversation_id)
            # Build LLM messages — convert user messages with file attachments to multimodal format
            # Filter out empty failed turns that pollute old chats (empty assistant with no tools/thinking)
            filtered_msgs = []
            for m in messages:
                if m.get("role") == "assistant" and not (m.get("content") or "").strip() and not m.get("thinking") and not (m.get("tool_calls") or []) and not ((m.get("metadata") or {}).get("blocks")):
                    print(f"[HISTORY] skipping empty failed assistant {m.get('id','')[:8]}")
                    continue
                filtered_msgs.append(m)
            messages = filtered_msgs
            # ── Regeneration context cut ─────────────────────────────
            # The model must answer the anchor prompt fresh: replay history only
            # up to (and including) that user message, and drop EVERY stored
            # version of the response being regenerated — otherwise the old
            # answer leaks into the context as if it were already spoken.
            if anchor_message_id:
                _anchor_idx = next(
                    (i for i, m in enumerate(messages) if m.get("id") == anchor_message_id),
                    -1,
                )
                if _anchor_idx >= 0:
                    messages = messages[:_anchor_idx + 1]
                if version_group:
                    messages = [m for m in messages if m.get("version_group") != version_group]
                print(f"[REGEN] history cut at anchor {anchor_message_id[:8]}: {len(messages)} msgs replayed")
            llm_messages = []
            for msg in messages:
                role = msg["role"]
                content = msg["content"]
                metadata = msg.get("metadata") or {}
                files = metadata.get("files", [])
                
                if role == "user" and files:
                    content_parts = []
                    if content.strip():
                        content_parts.append({"type": "text", "text": content})
                    for f in files:
                        f_type = f.get("type", "") or ""
                        f_url = f.get("url", "")
                        f_name = f.get("filename", "file")
                        if f_type.startswith("image/") and f_url:
                            # Convert image to base64 data URI for vision models
                            f_path = os.path.join(UPLOAD_DIR, os.path.basename(f_url))
                            if os.path.exists(f_path):
                                try:
                                    import base64
                                    with open(f_path, "rb") as img_f:
                                        b64_data = base64.b64encode(img_f.read()).decode()
                                    content_parts.append({
                                        "type": "image_url",
                                        "image_url": {"url": f"data:{f_type};base64,{b64_data}"}
                                    })
                                except Exception as e:
                                    print(f"[MULTIMODAL] Error reading image {f_path}: {e}")
                                    content_parts.append({"type": "text", "text": f"\n[Image: {f_name}]"})
                            else:
                                content_parts.append({"type": "text", "text": f"\n[Image file not found: {f_name}]"})
                        else:
                            # Try to read and include file content for text-based files and PDFs
                            f_path = os.path.join(UPLOAD_DIR, os.path.basename(f_url))
                            text_exts = {'.txt','.md','.json','.csv','.yaml','.yml','.xml','.log',
                                         '.py','.js','.ts','.jsx','.tsx','.html','.css','.scss','.less',
                                         '.sh','.bash','.zsh','.env','.ini','.cfg','.conf',
                                         '.sql','.r','.rb','.php','.go','.rs','.java','.kt','.swift',
                                         '.c','.cpp','.h','.hpp','.toml','.gradle','.makefile','.dockerfile',
                                         '.diff','.patch','.svg'}
                            file_ext = os.path.splitext(f_name)[1].lower()

                            if file_ext == '.pdf' and os.path.exists(f_path):
                                try:
                                    from pypdf import PdfReader
                                    pdf_content = []
                                    char_count = 0
                                    for page in PdfReader(f_path).pages:
                                        text = (page.extract_text() or '').strip()
                                        if text:
                                            remaining = 100000 - char_count
                                            if remaining <= 0:
                                                pdf_content.append('[...truncated...]')
                                                break
                                            pdf_content.append(text[:remaining])
                                            char_count += len(text[:remaining])
                                    pdf_text = '\n\n'.join(pdf_content)
                                    content_parts.append({
                                        "type": "text",
                                        "text": f"\n\nThe user attached the PDF '{f_name}'. Its content:\n\n{pdf_text}"
                                    })
                                except ImportError:
                                    content_parts.append({"type": "text", "text": f"\n[Attached PDF: {f_name}]"})
                                except Exception as e:
                                    print(f"[FILE] Error reading PDF {f_path}: {e}")
                                    content_parts.append({"type": "text", "text": f"\n[Attached PDF: {f_name}]"})

                            elif file_ext == '.docx' and os.path.exists(f_path):
                                try:
                                    from docx import Document
                                    paras = [p.text for p in Document(f_path).paragraphs if p.text]
                                    doc_text = '\n\n'.join(paras)[:100000]
                                    content_parts.append({
                                        "type": "text",
                                        "text": f"\n\nThe user attached the Word document '{f_name}'. Its content:\n\n{doc_text}"
                                    })
                                except ImportError:
                                    content_parts.append({"type": "text", "text": f"\n[Attached Word document: {f_name}]"})
                                except Exception as e:
                                    print(f"[FILE] Error reading DOCX {f_path}: {e}")
                                    content_parts.append({"type": "text", "text": f"\n[Attached Word document: {f_name}]"})

                            elif file_ext == '.xlsx' and os.path.exists(f_path):
                                try:
                                    from openpyxl import load_workbook
                                    wb = load_workbook(f_path, read_only=True, data_only=True)
                                    sheets_text = []
                                    char_count = 0
                                    for sheet_name in wb.sheetnames:
                                        ws = wb[sheet_name]
                                        rows = []
                                        max_rows = 60
                                        for i, row in enumerate(ws.iter_rows(values_only=True)):
                                            if i >= max_rows:
                                                rows.append(f'... ({ws.max_row or "?"} total rows)')
                                                break
                                            row_vals = [str(c) if c is not None else '' for c in row]
                                            rows.append(' | '.join(row_vals))
                                        sheet_text = f'--- Sheet: {sheet_name} ---\n' + '\n'.join(rows)
                                        remaining = 100000 - char_count
                                        if remaining <= 0:
                                            break
                                        sheets_text.append(sheet_text[:remaining])
                                        char_count += len(sheet_text[:remaining])
                                    wb.close()
                                    xlsx_text = '\n\n'.join(sheets_text)
                                    content_parts.append({
                                        "type": "text",
                                        "text": f"\n\nThe user attached the spreadsheet '{f_name}'. Its content:\n\n{xlsx_text}"
                                    })
                                except ImportError:
                                    content_parts.append({"type": "text", "text": f"\n[Attached spreadsheet: {f_name}]"})
                                except Exception as e:
                                    print(f"[FILE] Error reading XLSX {f_path}: {e}")
                                    content_parts.append({"type": "text", "text": f"\n[Attached spreadsheet: {f_name}]"})

                            elif file_ext in text_exts and os.path.exists(f_path):
                                try:
                                    with open(f_path, "r", encoding="utf-8", errors="replace") as text_f:
                                        file_content = text_f.read(100000)
                                    content_parts.append({
                                        "type": "text",
                                        "text": f"\n\nThe user attached the file '{f_name}'. Its content:\n\n```{file_ext.lstrip('.')}\n{file_content}\n```"
                                    })
                                except Exception as e:
                                    print(f"[FILE] Error reading {f_path}: {e}")
                                    content_parts.append({"type": "text", "text": f"\n[Attached file: {f_name}]"})

                            else:
                                content_parts.append({"type": "text", "text": f"\n[Attached file: {f_name}]"})
                    llm_messages.append({"role": role, "content": content_parts})
                else:
                    # ── KV-friendly history replay (tool-aware) ──────────
                    # For non-tool turns: single assistant message.
                    # For tool turns: emit assistant(with tool_calls + pre-content) →
                    # tool result(s) → assistant(post-content) to exactly match
                    # the live loop's token sequence (assistant->tool->assistant).
                    # IDs are replayed verbatim from DB for byte-exact prefix.
                    blocks = (metadata.get("blocks") or []) if isinstance(metadata, dict) else []
                    tool_blocks = [b for b in blocks if b.get("type") == "tool_call"] if role == "assistant" else []
                    has_tools = bool(role == "assistant" and msg.get("tool_calls") and tool_blocks)
                    if not has_tools:
                        llm_msg = {"role": role, "content": content}
                        if role == "assistant" and msg.get("thinking"):
                            llm_msg["reasoning_content"] = msg["thinking"]
                        llm_messages.append(llm_msg)
                    else:
                        # Split blocks into pre-tool vs post-tool content/thinking
                        first_tool_pos = next((i for i, b in enumerate(blocks) if b.get("type") == "tool_call"), len(blocks))
                        last_tool_pos = len(blocks) - 1 - next((i for i, b in enumerate(reversed(blocks)) if b.get("type") == "tool_call"), len(blocks))
                        pre_content = "".join(b.get("content", "") for b in blocks[:first_tool_pos] if b.get("type") == "content")
                        pre_thinking = "".join(b.get("content", "") for b in blocks[:first_tool_pos] if b.get("type") == "thinking")
                        # post_content/thinking is content after last tool_call; fallback to msg fields if empty (legacy rows)
                        post_parts = [b.get("content", "") for b in blocks[last_tool_pos+1:] if b.get("type") == "content"]
                        post_content = "".join(post_parts) if post_parts else (content or "")
                        post_thinking_parts = [b.get("content", "") for b in blocks[last_tool_pos+1:] if b.get("type") == "thinking"]
                        post_thinking = "".join(post_thinking_parts)
                        # Fallback for legacy rows where thinking was aggregated into column
                        if not pre_thinking and not post_thinking and msg.get("thinking"):
                            # Heuristic: if there was a tool call, thinking likely belongs before it; otherwise after
                            pre_thinking = msg["thinking"]
                            post_thinking = ""
                        # Build tool_calls with stored IDs (or deterministic fallback)
                        tcs = []
                        for idx, tc in enumerate(msg["tool_calls"] or []):
                            try:
                                args = tc.get("arguments") or {}
                                if not isinstance(args, dict):
                                    args = {}
                            except Exception:
                                args = {}
                            tid = tc.get("id") or tc.get("tool_call_id") or f"{tc.get('name','tool')}_{idx}"
                            # Prefer block's stored id if column id was missing (backward compat)
                            if idx < len(tool_blocks) and tool_blocks[idx].get("id"):
                                tid = tool_blocks[idx]["id"]
                            tcs.append({
                                "id": tid,
                                "type": "function",
                                "function": {
                                    "name": tc.get("name","unknown"),
                                    "arguments": __import__("json").dumps(args),
                                },
                            })
                        # First assistant: tool_calls + pre_content + pre_thinking
                        llm_msg = {"role": "assistant", "content": pre_content}
                        if pre_thinking:
                            llm_msg["reasoning_content"] = pre_thinking
                        if tcs:
                            llm_msg["tool_calls"] = tcs
                        llm_messages.append(llm_msg)
                        # Tool results + vision injection for replay
                        for idx, b in enumerate(tool_blocks):
                            tid = b.get("id")
                            if not tid and idx < len(tcs):
                                tid = tcs[idx]["id"]
                            else:
                                tid = tid or f"{b.get('name','tool')}_{idx}"
                            result = b.get("result")
                            try:
                                tool_result_str = __import__("json").dumps(result, default=str) if result is not None else "No result"
                            except Exception:
                                tool_result_str = str(result) if result is not None else "No result"
                            llm_messages.append({
                                "role": "tool",
                                "content": tool_result_str,
                                "tool_call_id": tid
                            })
                            # Replay images stored in block result (from screenshot / MCP ImageContent)
                            try:
                                _rimgs = (result or {}).get("images") if isinstance(result, dict) else None
                                if _rimgs:
                                    _rparts = [{"type": "text", "text": f"[Tool {b.get('name','tool')} returned {len(_rimgs)} image(s)]"}]
                                    for _rim in _rimgs:
                                        _rb64 = _rim.get("base64") or _rim.get("data") or ""
                                        _rmime = _rim.get("mime_type") or _rim.get("mimeType") or "image/png"
                                        if _rb64:
                                            _rparts.append({"type": "image_url", "image_url": {"url": f"data:{_rmime};base64,{_rb64}"}})
                                    if len(_rparts) > 1:
                                        llm_messages.append({"role": "user", "content": _rparts})
                            except Exception:
                                pass
                        if (post_content and post_content.strip()) or post_thinking:
                            post_msg = {"role": "assistant", "content": post_content}
                            if post_thinking:
                                post_msg["reasoning_content"] = post_thinking
                            llm_messages.append(post_msg)

            # ── Auto-recall: keyword-search long-term memory for the CURRENT user
            # message and append hits to the user turn itself (never the system
            # prompt → KV prefix stays hot, DB row untouched, prompt copy only).
            # This is what makes the memory store actually reach answers without
            # the model having to think of calling memory_search.
            try:
                from database.memory_crud import get_recall_for_query
                _lu_idx = next((i for i in range(len(llm_messages) - 1, -1, -1)
                                if llm_messages[i].get("role") == "user"), None)
                if _lu_idx is not None:
                    _c = llm_messages[_lu_idx].get("content")
                    if isinstance(_c, str):
                        _q = _c
                    elif isinstance(_c, list):
                        _q = " ".join(p.get("text", "") for p in _c if isinstance(p, dict) and p.get("type") == "text")
                    else:
                        _q = ""
                    if _q.strip():
                        _recall = await get_recall_for_query(db, _q, top_k=5)
                        if _recall:
                            if isinstance(_c, str):
                                llm_messages[_lu_idx]["content"] = _c + "\n\n" + _recall
                            elif isinstance(_c, list):
                                llm_messages[_lu_idx]["content"] = list(_c) + [{"type": "text", "text": "\n\n" + _recall}]
                            print(f"[MEMORY] auto-recalled {_recall.count(chr(10)) - 3} entr(ies) into user turn")
            except Exception as e:
                print(f"[MEMORY] auto-recall failed: {e}")
            
            # Prepend system prompt to messages
            if system_prompt_content:
                llm_messages.insert(0, {"role": "system", "content": system_prompt_content})

            # ponytail: truncate huge history for old chats (99 msgs → 400 on free models)
            # Free models fail even with 21 msgs; keep last 10 for -free, 40 otherwise
            is_free = bool(model and "-free" in model)
            limit = 10 if is_free else 40
            thresh = 15 if is_free else 50
            if len(llm_messages) > thresh:
                sys_msg = llm_messages[0] if llm_messages and llm_messages[0].get("role") == "system" else None
                rest = llm_messages[1:] if sys_msg else llm_messages
                keep = rest[-limit:]
                llm_messages = ([sys_msg] + keep) if sys_msg else keep
                print(f"[HISTORY] truncated {len(rest)+ (1 if sys_msg else 0)} → {len(llm_messages)} for model {model} (free={is_free})")

            tool_calls_history = []

            # Track message blocks for sequential display (content, thinking, tool calls)
            message_blocks = []

            # ── Placeholder for crash/stop/approval safety ────────────
            # The previous turn is already committed in its own handler.
            # This placeholder guarantees the *current* turn is never lost:
            # created empty and updated incrementally as chunks arrive,
            # on every tool event, and immediately on approval. A restart or
            # Stop therefore finds a row instead of nothing.
            placeholder_id: Optional[str] = None
            _last_persist = 0.0
            try:
                _ph = await add_message(
                    db, conversation_id, "assistant", "",
                    blocks=[], extra_metadata={"model": model} if model else {},
                    version=version, version_group=version_group,
                    turn_index=turn_index,
                )
                placeholder_id = _ph["id"]
                await db.commit()
                print(f"[PERSIST] placeholder {placeholder_id[:8]} for {request_id[:8]}")
            except Exception as _e:
                print(f"[PERSIST] placeholder create failed: {_e}")
                placeholder_id = None

            async def _flush_placeholder(force: bool = False):
                nonlocal _last_persist
                if not placeholder_id:
                    return
                now = _time.monotonic()
                if not force and (now - _last_persist < 0.8):
                    return
                _last_persist = now
                try:
                    _cons = []
                    for _b in message_blocks:
                        _bt = _b.get("type")
                        if _bt in ("content", "thinking"):
                            if _cons and _cons[-1].get("type") == _bt:
                                _cons[-1]["content"] = _cons[-1].get("content","") + _b.get("content","")
                            else:
                                _cons.append(dict(_b))
                        else:
                            _cons.append(dict(_b))
                    _full_content = "".join(b.get("content","") for b in _cons if b.get("type")=="content")
                    _full_thinking = "".join(b.get("content","") for b in _cons if b.get("type")=="thinking")
                    _extra = {"model": model} if model else {}
                    _fm = _aggregate_turn_metrics(turn_metrics)
                    if _fm:
                        _extra["metrics"] = _fm
                    await update_assistant_message_full(
                        db, placeholder_id,
                        content=_full_content,
                        thinking=_full_thinking,
                        blocks=_cons if _cons else message_blocks,
                        extra_metadata=_extra or None,
                    )
                except Exception as _e:
                    print(f"[PERSIST] flush failed: {_e}")

            # Get MCP tools for LLM function calling
            mcp_tools = []
            if mcp_manager:
                # Get MCP tools, excluding disabled ones to save tokens in the LLM prompt.
                mcp_tools = await mcp_manager.list_all_tools(include_disabled=False)
                if mcp_tools:
                    print(f"[TOOLS] {len(mcp_tools)} MCP tools available for LLM")

            # ── Filter tools by agent binding ──────────────────────────
            # Agent can restrict which MCP servers and custom tools are sent to the LLM.
            # Empty list = allow everything (backward compatible).
            # Non-empty list = only allow those specific servers/tools.
            enabled_mcp_servers = []
            if agent_config and agent_config.get("enabled_mcp_servers"):
                enabled_mcp_servers = agent_config["enabled_mcp_servers"]
            if enabled_mcp_servers:
                allowed_servers = set(enabled_mcp_servers)
                # One-off session overrides (user manually enabled a server the
                # agent doesn't allow): union them in for this request only.
                if override_servers:
                    allowed_servers |= set(override_servers)
                    print(f"[TOOLS] Session overrides: {sorted(set(override_servers))}")
                mcp_tools = [t for t in mcp_tools if t.get("server") in allowed_servers]
                print(f"[TOOLS] Filtered to {len(mcp_tools)} MCP tools from servers: {allowed_servers}")

            # ── Filter custom tools by agent's enabled_tools ──────────────
            # Custom tools are controlled differently:
            #   - query_documents (RAG) → controlled by enable_rag flag
            #   - generate_speech (TTS) → controlled by exclude_tools param
            # If enabled_tools is configured, exclude tools not in the list.
            exclude_tools = []
            effective_enable_rag = enable_rag
            if agent_config and agent_config.get("enabled_tools"):
                enabled_custom = set(agent_config["enabled_tools"])
                if "query_documents" not in enabled_custom:
                    effective_enable_rag = False
                for custom_tool in custom_tool_names():
                    if custom_tool not in enabled_custom:
                        exclude_tools.append(custom_tool)
            elif agent_config is not None:
                # No specific tools restriction — use agent's enable_rag setting
                if agent_config.get("enable_rag"):
                    effective_enable_rag = True

            # Get tool definitions with agent-based filtering
            all_tools = tool_executor.get_tool_definitions(
                exclude_tools=exclude_tools,
                mcp_tools=mcp_tools,
                enable_rag=effective_enable_rag
            )

            # Debug logging - show exactly what tools are being sent
            print(f"[TOOLS] MCP tools discovered: {len(mcp_tools)}")
            print(f"[TOOLS] Total tools sent to LLM: {len(all_tools)}")
            print(f"[TOOLS] enable_rag={effective_enable_rag}")
            for tool in all_tools:
                tool_name = tool.get("function", {}).get("name", "unknown")
                tool_server = tool.get("server", "builtin")
                print(f"[TOOLS]   - {tool_name} ({'MCP: ' + tool_server if tool_server != 'builtin' else 'builtin'})")
            if mcp_tools:
                for tool in mcp_tools:
                    print(f"  - MCP: {tool['name']} from {tool['server']}")

            # Context transparency event (Phase 6) — what the model sees this turn.
            try:
                yield f"data: {json.dumps({'type': 'context_info', 'context': {
                    'model': model or 'default',
                    'thinking_mode': thinking_mode or 'auto',
                    'system_prompt': system_prompt_content,
                    'message_count': len(llm_messages),
                    'tool_count': len(all_tools),
                    'tools': [t.get('function', {}).get('name', '?') for t in all_tools],
                }})}\n\n"
            except Exception as e:
                print(f"[CONTEXT] event failed: {e}")
            
            # Save original prefix for KV-friendly autos (Phase A) — llm_messages will be mutated in the loop
            original_llm_messages = list(llm_messages)
            # KV-preserving steer: set when frontend queues a steer; we break
            # gracefully at the next chunk boundary so the llama.cpp slot's KV
            # is retained (clean aclose) and the partial is flushed.
            steered = False
            # Main conversation loop - handles multiple tool calls with content in between
            max_tool_iterations = 35  # Prevent infinite loops
            tool_iteration = 0
            while tool_iteration < max_tool_iterations:
                # Steer can arrive between tool iterations (e.g. during tool run)
                if await _is_steer_pending(conversation_id):
                    print(f"[STEER] interrupt before iteration {tool_iteration+1} for {conversation_id[:8]}")
                    steered = True
                    try:
                        await _flush_placeholder(force=True)
                    except Exception:
                        pass
                    try:
                        yield f"data: {json.dumps({'type': 'steer_ack', 'conversation_id': conversation_id})}\n\n"
                    except Exception:
                        pass
                    break
                tool_iteration += 1
                print(f"[DEBUG] Conversation loop iteration {tool_iteration}")

                # Stream LLM response
                assistant_message, thinking_content = "", ""
                pending_tool_calls = []

                # Two-phase stall watchdog: generous initial timeout for prompt
                # processing (long contexts can take 80s+), shorter timeout for
                # inter-chunk stalls. Continues indefinitely while chunks arrive.
                try:
                    async for chunk in _stream_with_stall_timeout(
                        llm_client.stream_chat(llm_messages, model=model, tools=all_tools,
                                               base_url=provider_base_url, api_key=provider_api_key,
                                               thinking_mode=thinking_mode,
                                               temperature=_agent_temperature,
                                               max_tokens=_agent_max_tokens,
                                               top_k=_agent_top_k),
                        initial_timeout=600,
                        stall_timeout=60
                    ):
                        try:
                            chunk_type = chunk.get("type")

                            if chunk_type == "content":
                                content = chunk.get("content", "")
                                assistant_message += content
                                if content:
                                    message_blocks.append({
                                        "type": "content",
                                        "content": content
                                    })
                                yield f"data: {json.dumps({'type': 'content', 'content': content})}\n\n"
                                # incremental persist (throttled) so Stop/restart never loses the whole response
                                try:
                                    await _flush_placeholder()
                                except Exception:
                                    pass
                            elif chunk_type == "thinking":
                                thinking = chunk.get("content", "")
                                thinking_content += thinking
                                if thinking:
                                    if message_blocks and message_blocks[-1].get('type') == 'thinking':
                                        message_blocks[-1]['content'] += thinking
                                    else:
                                        message_blocks.append({
                                            "type": "thinking",
                                            "content": thinking
                                        })
                                yield f"data: {json.dumps({'type': 'thinking', 'content': thinking})}\n\n"
                                try:
                                    await _flush_placeholder()
                                except Exception:
                                    pass
                            elif chunk_type == "tool_call":
                                print(f"[DEBUG] Tool call chunk received")
                                tc_data = chunk.get("tool_call")
                                print(f"[DEBUG] tc_data: {tc_data}")

                                if tc_data:
                                    tool_name = tc_data.get('name')
                                    tool_args = tc_data.get('arguments')

                                    if tool_name:
                                        call_key = f"{tool_name}_{len(pending_tool_calls)}"
                                        pending_tool_calls.append({
                                            "name": tool_name,
                                            "arguments": tool_args if isinstance(tool_args, dict) else {},
                                            "status": "pending",
                                            "result": None,
                                            "progress_history": [],
                                            "key": call_key
                                        })
                                        print(f"[DEBUG] Pending tool call #{len(pending_tool_calls)-1}: {tool_name}, args: {pending_tool_calls[-1]['arguments']}")

                            elif chunk_type == "tool_error":
                                # Premature EOS from MTP — tool call was started but never completed
                                tool_name = chunk.get("tool", "unknown")
                                error_msg = chunk.get("error", "Incomplete tool call")
                                print(f"[DEBUG] Incomplete tool call from stream: {tool_name}: {error_msg}")
                                # Add an error block so the user sees what happened
                                message_blocks.append({
                                    "type": "tool_call",
                                    "name": tool_name,
                                    "arguments": {},
                                    "status": "error",
                                    "result": {"error": error_msg},
                                    "progress_history": []
                                })
                                yield f"data: {json.dumps({'type': 'tool_error', 'tool': tool_name, 'error': error_msg})}\n\n"
                                try:
                                    await _flush_placeholder(force=True)
                                except Exception:
                                    pass

                            elif chunk_type == "metrics":
                                m = chunk.get("metrics") or {}
                                turn_metrics.append(m)
                                # Forward live so UI can show streaming speed without waiting for DB
                                yield f"data: {json.dumps({'type': 'metrics', 'metrics': m})}\n\n"

                            elif chunk_type == "error":
                                # surface it instead of silently ending with an empty reply.
                                error_msg = chunk.get("error", "LLM request failed")
                                print(f"[LLM ERROR] {error_msg}")
                                yield f"data: {json.dumps({'type': 'error', 'error': error_msg})}\n\n"
                                break

                        except Exception as e:
                            print(f"[DEBUG] Error processing chunk: {e}")
                            print(f"[DEBUG] Chunk data: {chunk}")
                            import traceback
                            traceback.print_exc()
                        await asyncio.sleep(0)
                        # KV-preserving graceful steer: check at chunk boundary
                        if await _is_steer_pending(conversation_id):
                            print(f"[STEER] graceful interrupt at chunk boundary for {conversation_id[:8]} ({len(assistant_message)} chars)")
                            steered = True
                            try:
                                await _flush_placeholder(force=True)
                            except Exception:
                                pass
                            try:
                                yield f"data: {json.dumps({'type': 'steer_ack', 'conversation_id': conversation_id})}\n\n"
                            except Exception:
                                pass
                            break
                except asyncio.TimeoutError:
                    # Safety net — _stream_with_stall_timeout already catches this,
                    # but handle it here to prevent unhandled exception propagation.
                    print(f"[WATCHDOG] Unhandled timeout in stream processing for request {request_id}")
                    # Fall through to save partial response and yield done

                # Steered: skip tool execution — save partial and exit quickly
                if steered:
                    print(f"[STEER] skipping tool execution for {conversation_id[:8]} due to steer")
                    break
                # If we have pending tool calls, execute them and continue the loop
                if pending_tool_calls:
                    # the response to a preceding assistant message carrying tool_calls.
                    # Some backends (llama.cpp proxying to strict upstream providers)
                    # reject the request with HTTP 400 if this pairing is missing.
                    assistant_tool_calls = []
                    for i, tc in enumerate(pending_tool_calls):
                        assistant_tool_calls.append({
                            "id": f"{tc['name']}_{i}",
                            "type": "function",
                            "function": {
                                "name": tc["name"],
                                "arguments": json.dumps(tc["arguments"]) if isinstance(tc["arguments"], dict) else str(tc.get("arguments") or "{}"),
                            },
                        })
                    assistant_tc_msg = {"role": "assistant", "tool_calls": assistant_tool_calls}
                    if assistant_message.strip():
                        assistant_tc_msg["content"] = assistant_message
                    # Strict reasoning providers require the original reasoning
                    # (thinking mode) to be echoed back with the assistant message.
                    if thinking_content:
                        assistant_tc_msg["reasoning_content"] = thinking_content
                    llm_messages.append(assistant_tc_msg)

                    for i, pending_tool_call in enumerate(pending_tool_calls):
                        # Steer can arrive while previous tool was running
                        if await _is_steer_pending(conversation_id):
                            print(f"[STEER] interrupt during tool loop for {conversation_id[:8]} before tool {pending_tool_call['name']}")
                            steered = True
                            break
                        print(f"Executing tool {i+1}/{len(pending_tool_calls)}: {pending_tool_call['name']}")
                        # Send tool call start event
                        yield f"data: {json.dumps({'type': 'tool_call_start', 'tool': pending_tool_call['name'], 'args': pending_tool_call['arguments']})}\n\n"
                        tool_calls_history.append(pending_tool_call)

                        # Execute the tool
                        tool_result = None
                        async for progress_event in tool_executor.execute_tool(
                            pending_tool_call['name'],
                            pending_tool_call['arguments'],
                            request_id,
                            call_key=pending_tool_call.get('key'),
                            conversation_id=conversation_id,
                            skill_allowlist=(agent_config or {}).get("enabled_skills") or None
                        ):
                            # Forward the progress event
                            yield f"data: {json.dumps(progress_event)}\n\n"
                            if progress_event.get("type") == "tool_approval_required":
                                # Persist approval state immediately so refresh/disconnect can resume
                                pending_tool_call["status"] = "approval"
                                pending_tool_call["command"] = progress_event.get("command", "")
                                pending_tool_call["working_dir"] = progress_event.get("working_dir", "")
                                pending_tool_call["approval_key"] = progress_event.get("approval_key")
                                pending_tool_call["approval_reason"] = progress_event.get("reason", "")
                                # Create provisional block for save-on-cancel
                                provisional = {
                                    "type": "tool_call",
                                    "id": f"{pending_tool_call['name']}_{i}",
                                    "name": pending_tool_call['name'],
                                    "arguments": pending_tool_call['arguments'],
                                    "status": "approval",
                                    "command": progress_event.get("command", ""),
                                    "working_dir": progress_event.get("working_dir", ""),
                                    "approval_key": progress_event.get("approval_key"),
                                    "approval_reason": progress_event.get("reason", ""),
                                    "result": None,
                                    "progress_history": pending_tool_call["progress_history"][:],
                                }
                                # Append or update provisional in message_blocks
                                # (so CancelledError save includes it)
                                if not any(b.get("type")=="tool_call" and b.get("name")==pending_tool_call['name'] and b.get("status")=="approval" for b in message_blocks):
                                    message_blocks.append(provisional)
                                    try:
                                        await _flush_placeholder(force=True)
                                    except Exception:
                                        pass
                            elif progress_event.get("type") == "tool_progress":
                                pending_tool_call["status"] = progress_event.get("status", "running")
                                pending_tool_call["progress"] = progress_event.get("progress", 0)
                                pending_tool_call["progress_history"].append(progress_event)
                                # Check for result in various possible locations
                                # For MCP tools: result is in 'content' field (parsed CallToolResult)
                                # For custom tools: result might be in 'result' field
                                result = progress_event.get("result") or progress_event.get("content")
                                if result:
                                    pending_tool_call["result"] = result
                                    pending_tool_call["status"] = "completed"
                                    tool_result = result
                            elif progress_event.get("type") == "tool_error":
                                pending_tool_call["status"] = "error"
                                pending_tool_call["result"] = {"error": progress_event.get("error")}
                                tool_result = {"error": progress_event.get("error")}
                                # If this was an approval timeout/deny, update provisional block if present
                                for b in message_blocks:
                                    if b.get("type")=="tool_call" and b.get("status")=="approval" and b.get("name")==pending_tool_call['name']:
                                        b["status"] = "error"
                                        b["result"] = {"error": progress_event.get("error")}
                                        break

                        # Add tool call block to message blocks for sequential display
                        # If a provisional approval block was already appended (for save-on-cancel),
                        # update it in place instead of duplicating.
                        provisional_idx = None
                        for idx, b in enumerate(message_blocks):
                            if b.get("type")=="tool_call" and b.get("name")==pending_tool_call['name'] and b.get("status")=="approval":
                                provisional_idx = idx
                                break
                        tool_call_block = {
                            "type": "tool_call",
                            "id": f"{pending_tool_call['name']}_{i}",
                            "name": pending_tool_call['name'],
                            "arguments": pending_tool_call['arguments'],
                            "status": pending_tool_call['status'],
                            "result": pending_tool_call['result'],
                            "progress_history": pending_tool_call['progress_history']
                        }
                        # Preserve command/approval fields if this was an approval flow
                        if pending_tool_call.get("command"):
                            tool_call_block["command"] = pending_tool_call["command"]
                            tool_call_block["working_dir"] = pending_tool_call.get("working_dir", "")
                            tool_call_block["approval_key"] = pending_tool_call.get("approval_key")
                            tool_call_block["approval_reason"] = pending_tool_call.get("approval_reason", "")
                        # Extract sources from result for bottom-of-chat display
                        if pending_tool_call['result'] and isinstance(pending_tool_call['result'], dict):
                            sources = pending_tool_call['result'].get('sources', [])
                            if sources:
                                tool_call_block['sources'] = sources
                        if provisional_idx is not None:
                            # Replace provisional with final status (error/completed)
                            message_blocks[provisional_idx] = tool_call_block
                        else:
                            message_blocks.append(tool_call_block)
                        try:
                            await _flush_placeholder(force=True)
                        except Exception:
                            pass

                        # Add tool result to conversation for LLM to continue
                        # Format for llama.cpp: role=tool with content as string
                        tool_result_str = json.dumps(tool_result, default=str) if tool_result else "No result"
                        print(f"[DEBUG] Tool result {i+1} preview: {tool_result_str[:500] if tool_result_str else 'None'}...")
                        llm_messages.append({
                            "role": "tool",
                            "content": tool_result_str,
                            "tool_call_id": f"{pending_tool_call['name']}_{i}"
                        })
                        # Vision: if tool returned images (MCP ImageContent or screenshot file path), inject as user image message
                        try:
                            _imgs = (tool_result or {}).get("images") if isinstance(tool_result, dict) else None
                            if _imgs:
                                _parts = [{"type": "text", "text": f"[Tool {pending_tool_call['name']} returned {len(_imgs)} image(s)]"}]
                                for _im in _imgs:
                                    _b64 = _im.get("base64") or _im.get("data") or ""
                                    _mime = _im.get("mime_type") or _im.get("mimeType") or "image/png"
                                    if _b64:
                                        _parts.append({"type": "image_url", "image_url": {"url": f"data:{_mime};base64,{_b64}"}})
                                if len(_parts) > 1:
                                    llm_messages.append({"role": "user", "content": _parts})
                                    print(f"[VISION] Injected {len(_imgs)} image(s) from tool {pending_tool_call['name']} into llm_messages")
                        except Exception as _ve:
                            print(f"[VISION] inject failed: {_ve}")
                    print(f"[DEBUG] All {len(pending_tool_calls)} tools executed, continuing conversation with results")
                    if steered:
                        print(f"[STEER] tools interrupted by steer — exiting loop for {conversation_id[:8]}")
                        break
                    # Record skill usage (Phase 4) for the improvement loop.
                    try:
                        from database.skill_crud import record_skill_run
                        for tc in pending_tool_calls:
                            if tc.get("name") == "load_skill":
                                skill_name = str((tc.get("arguments") or {}).get("name", "")).strip()
                                if skill_name:
                                    await record_skill_run(
                                        db, skill_name, conversation_id,
                                        success=tc.get("status") not in ("error",)
                                    )
                        await db.commit()
                    except Exception as e:
                        print(f"[SKILLS] run-log failed: {e}")
                    # Continue the while loop to get LLM's response to the tool result
                    # (LLM may respond with content, thinking, or another tool call)

                else:
                    # No tool call, conversation is complete
                    print(f"[DEBUG] No pending tool call, conversation complete")
                    break
            
            # Log total messages sent to LLM after all iterations
            print(f"[DEBUG] Total messages in llm_messages after {tool_iteration} iterations: {len(llm_messages)}")
            
            # Aggregate performance metrics across all LLM calls in this turn
            final_metrics = _aggregate_turn_metrics(turn_metrics)
            if final_metrics and len(turn_metrics) > 1:
                try:
                    yield f"data: {json.dumps({'type': 'metrics', 'metrics': final_metrics, 'aggregated': len(turn_metrics) > 1})}\n\n"
                except Exception as _e:
                    print(f"[METRICS] emit aggregated failed: {_e}")
            
            # Save assistant message — update placeholder if it exists (crash-safe),
            # otherwise create new. The placeholder was committed at stream start
            # and updated incrementally, so the previous turn is already safe and
            # the current partial is never lost on Stop/restart/approval.
            if placeholder_id:
                # Consolidate like _save_assistant_message, then update in place
                _cons2 = []
                for _b in message_blocks:
                    _bt = _b.get("type")
                    if _bt in ("content", "thinking"):
                        if _cons2 and _cons2[-1].get("type") == _bt:
                            _cons2[-1]["content"] = _cons2[-1].get("content","") + _b.get("content","")
                        else:
                            _cons2.append(dict(_b))
                    else:
                        _cons2.append(dict(_b))
                _full_c = "".join(b.get("content","") for b in _cons2 if b.get("type")=="content")
                if not _full_c:
                    _full_c = assistant_message
                _full_t = "".join(b.get("content","") for b in _cons2 if b.get("type")=="thinking")
                if not _full_t:
                    _full_t = thinking_content
                _extra2: Dict = {}
                if model:
                    _extra2["model"] = model
                if final_metrics:
                    _extra2["metrics"] = final_metrics
                # If nothing to save and placeholder is still empty, keep it (still counts as saved for done)
                if not (_full_c.strip() or _full_t.strip() or _cons2):
                    # Keep the empty placeholder as the saved row (so done has an id)
                    from sqlalchemy import select as _sel2
                    from backend.database.models import Message as _M2
                    _r2 = await db.execute(_sel2(_M2).where(_M2.id == placeholder_id))
                    _row2 = _r2.scalar_one_or_none()
                    if _row2 is not None and final_metrics:
                        _meta2 = dict(_row2.extra_metadata or {})
                        if model:
                            _meta2["model"] = model
                        _meta2["metrics"] = final_metrics
                        _row2.extra_metadata = _meta2
                        await db.commit()
                    assistant_saved = {"id": placeholder_id, "version_group": version_group}
                    saved_msg_id = placeholder_id
                    saved_version_group = version_group
                else:
                    assistant_saved = await update_assistant_message_full(
                        db, placeholder_id,
                        content=_full_c,
                        thinking=_full_t,
                        blocks=_cons2 if _cons2 else message_blocks,
                        extra_metadata=_extra2 if _extra2 else None,
                    )
                    # update_assistant_message_full already committed
                    saved_msg_id = placeholder_id
                    saved_version_group = version_group
                    # keep the dict shape for later activity_blocks handling
                    if assistant_saved is None:
                        # placeholder was deleted? fallback to create
                        assistant_saved = await _save_assistant_message(
                            db, conversation_id, assistant_message, thinking_content,
                            message_blocks, model, version, version_group,
                            turn_index=turn_index,
                            metrics=final_metrics
                        )
                        saved_msg_id = assistant_saved["id"] if isinstance(assistant_saved, dict) else placeholder_id
                        saved_version_group = assistant_saved.get("version_group") if isinstance(assistant_saved, dict) else version_group
                        placeholder_id = saved_msg_id
            else:
                assistant_saved = await _save_assistant_message(
                    db, conversation_id, assistant_message, thinking_content,
                    message_blocks, model, version, version_group,
                    turn_index=turn_index,
                    metrics=final_metrics
                )
                saved_msg_id = assistant_saved["id"] if isinstance(assistant_saved, dict) else None
                saved_version_group = assistant_saved.get("version_group") if isinstance(assistant_saved, dict) else version_group
                placeholder_id = saved_msg_id
            # ── Steered fast-path: skip heavy post-actions so SSE closes fast
            # and next turn's prefill can reuse the full prefix KV.
            if steered:
                await _consume_steer_pending(conversation_id)
                print(f"[STEER] fast done for {conversation_id[:8]} — saved {len(assistant_message)} chars, {len(message_blocks)} blocks, skipping autos")
                # Persist a minimal steer marker so the cut is visible on reload
                try:
                    from sqlalchemy import select as _select_s
                    from backend.database.models import Message as _MsgS
                    _rs = await db.execute(_select_s(_MsgS).where(_MsgS.id == saved_msg_id))
                    _row_s = _rs.scalar_one_or_none()
                    if _row_s is not None:
                        _meta_s = dict(_row_s.extra_metadata or {})
                        _existing_s = list(_meta_s.get("blocks") or [])
                        _meta_s["blocks"] = _existing_s + [{"type": "auto_action", "action": "steer", "status": "completed", "detail": {"chars": len(assistant_message), "blocks": len(message_blocks)}, "ts": int(_time.time()*1000)}]
                        _row_s.extra_metadata = _meta_s
                        await db.commit()
                except Exception as _e_s:
                    print(f"[STEER] marker persist failed: {_e_s}")
                try:
                    yield f"data: {json.dumps({'type': 'done', 'message_id': saved_msg_id, 'version_group': saved_version_group, 'version': version, 'steered': True})}\n\n"
                except Exception:
                    pass
                return
            # and streamed live so the thread shows "what the model did after answering".
            import time as _aa_time
            activity_blocks: list = []
            def _aa_block(action: str, status: str, detail: dict | None = None):
                return {
                    "type": "auto_action",
                    "action": action,
                    "status": status,
                    "detail": detail or {},
                    "ts": int(_aa_time.time() * 1000),
                }

            # Finalize open job runs for this conversation (Phase 5).
            if assistant_saved:
                try:
                    from database.job_crud import list_job_runs, finish_job_run
                    import os as _os
                    from settings import OUTPUTS_DIR
                    runs = await list_job_runs(db, limit=200)
                    open_runs = [r for r in runs
                                 if r["conversation_id"] == conversation_id and r["status"] == "running"]
                    if open_runs:
                        # stream start
                        yield f"data: {json.dumps({'type': 'auto_action', 'action': 'jobs', 'status': 'running', 'detail': {'count': len(open_runs)}})}\n\n"
                        jobs_dir = _os.path.join(OUTPUTS_DIR, "jobs")
                        _os.makedirs(jobs_dir, exist_ok=True)
                        for run in open_runs:
                            if assistant_message.strip():
                                output_path = _os.path.join(jobs_dir, f"{run['id']}.md")
                                try:
                                    with open(output_path, "w", encoding="utf-8") as f:
                                        f.write(f"# Job: {run['job_name']}\n\n{assistant_message}")
                                except Exception as e:
                                    print(f"[JOBS] output write failed: {e}")
                                    output_path = None
                                await finish_job_run(db, run["id"], "completed", output_path=output_path)
                            else:
                                await finish_job_run(db, run["id"], "failed",
                                                     error="No assistant output produced")
                        await db.commit()
                        print(f"[JOBS] finalized {len(open_runs)} run(s) for conversation {conversation_id[:8]}")
                        blk = _aa_block("jobs", "completed", {"count": len(open_runs), "run_ids": [r["id"] for r in open_runs]})
                        activity_blocks.append(blk)
                        yield f"data: {json.dumps({'type': 'auto_action', 'action': 'jobs', 'status': 'completed', 'detail': blk['detail']})}\n\n"
                except Exception as e:
                    print(f"[JOBS] finalize failed: {e}")
                    blk = _aa_block("jobs", "error", {"reason": str(e)[:400]})
                    activity_blocks.append(blk)
                    yield f"data: {json.dumps({'type': 'auto_action', 'action': 'jobs', 'status': 'error', 'detail': blk['detail']})}\n\n"

            # Auto memory extraction (Phase 2) — insight-based, KV-friendly
            if assistant_saved:
                # stream start immediately so UI shows spinner even for skipped cadence
                yield f"data: {json.dumps({'type': 'auto_action', 'action': 'memory', 'status': 'running'})}\n\n"
                mem_res = None
                try:
                    mem_res = await asyncio.wait_for(
                        _extract_memory_from_exchange(
                            db, conversation_id, llm_client,
                            agent_id=conversation.agent_id if conversation else None,
                            model=model,
                            base_url=provider_base_url, api_key=provider_api_key,
                            llm_messages=original_llm_messages, assistant_message=assistant_message,
                            thinking_content=thinking_content,
                            tools=all_tools, thinking_mode=thinking_mode, message_blocks=message_blocks
                        ),
                        timeout=300
                    )
                except asyncio.TimeoutError:
                    print("[MEMORY] extraction timed out")
                    mem_res = {"action": "memory", "status": "error", "detail": {"reason": "timeout (90s)"}}
                except Exception as e:
                    print(f"[MEMORY] extraction error: {e}")
                    mem_res = {"action": "memory", "status": "error", "detail": {"reason": str(e)[:400]}}
                # Normalize result
                if not isinstance(mem_res, dict) or "status" not in mem_res:
                    mem_res = {"action": "memory", "status": "skipped", "detail": {"reason": "no result"}}
                blk = _aa_block("memory", mem_res.get("status", "skipped"), mem_res.get("detail", {}))
                activity_blocks.append(blk)
                yield f"data: {json.dumps({'type': 'auto_action', 'action': 'memory', 'status': blk['status'], 'detail': blk['detail']})}\n\n"

                # Self-improvement reflection (Phase 4) — proposes skill drafts (insight-based, KV-friendly)
                yield f"data: {json.dumps({'type': 'auto_action', 'action': 'skill', 'status': 'running'})}\n\n"
                skill_res = None
                try:
                    skill_res = await asyncio.wait_for(
                        _maybe_reflect_and_propose_skill(db, conversation_id, llm_client, model=model,
                                                         base_url=provider_base_url, api_key=provider_api_key,
                                                         llm_messages=original_llm_messages, assistant_message=assistant_message,
                                                         thinking_content=thinking_content, tools=all_tools,
                                                         thinking_mode=thinking_mode, message_blocks=message_blocks),
                        timeout=300
                    )
                except asyncio.TimeoutError:
                    print("[SKILLS] reflection timed out")
                    skill_res = {"action": "skill", "status": "error", "detail": {"reason": "timeout (90s)"}}
                except Exception as e:
                    print(f"[SKILLS] reflection error: {e}")
                    skill_res = {"action": "skill", "status": "error", "detail": {"reason": str(e)[:400]}}
                if not isinstance(skill_res, dict) or "status" not in skill_res:
                    skill_res = {"action": "skill", "status": "skipped", "detail": {"reason": "no result"}}
                blk2 = _aa_block("skill", skill_res.get("status", "skipped"), skill_res.get("detail", {}))
                activity_blocks.append(blk2)
                yield f"data: {json.dumps({'type': 'auto_action', 'action': 'skill', 'status': blk2['status'], 'detail': blk2['detail']})}\n\n"

            # Generate title using model (reuses KV cache via cache_prompt: true)
            # Skip if the conversation already has a meaningful title (not the default)
            existing_title = conversation.title if conversation else None
            if assistant_saved and (not existing_title or existing_title == 'New Chat'):
                async with get_db() as title_db:
                    msgs = await get_conversation_messages(title_db, conversation_id)
                    user_count = len([m for m in msgs if m["role"] == "user"])
                    assistant_count = len([m for m in msgs if m["role"] == "assistant"])

                should_title = (user_count == 1 and assistant_count == 1)
                if should_title:
                    yield f"data: {json.dumps({'type': 'auto_action', 'action': 'title', 'status': 'running'})}\n\n"
                    title = ""
                    try:
                        title = await _generate_title_with_model(
                            original_llm_messages, assistant_message, llm_client, tools=all_tools, model=model,
                            base_url=provider_base_url, api_key=provider_api_key,
                            thinking_content=thinking_content, thinking_mode=thinking_mode,
                            message_blocks=message_blocks
                        )
                    except Exception as e:
                        print(f"[TITLE] generation failed: {e}")
                        blk3 = _aa_block("title", "error", {"reason": str(e)[:400]})
                        activity_blocks.append(blk3)
                        yield f"data: {json.dumps({'type': 'auto_action', 'action': 'title', 'status': 'error', 'detail': blk3['detail']})}\n\n"
                        title = ""
                    else:
                        if title:
                            try:
                                await update_conversation_title(db, conversation_id, title)
                                await db.commit()
                            except Exception as e:
                                print(f"[TITLE] db update failed: {e}")
                            yield f"data: {json.dumps({'type': 'title_update', 'title': title})}\n\n"
                            blk3 = _aa_block("title", "completed", {"title": title})
                            activity_blocks.append(blk3)
                            yield f"data: {json.dumps({'type': 'auto_action', 'action': 'title', 'status': 'completed', 'detail': blk3['detail']})}\n\n"
                        else:
                            blk3 = _aa_block("title", "skipped", {"reason": "model returned empty title"})
                            activity_blocks.append(blk3)
                            yield f"data: {json.dumps({'type': 'auto_action', 'action': 'title', 'status': 'skipped', 'detail': blk3['detail']})}\n\n"
                else:
                    blk3 = _aa_block("title", "skipped", {"reason": f"not first turn (u:{user_count} a:{assistant_count}) or custom title"})
                    activity_blocks.append(blk3)
                    yield f"data: {json.dumps({'type': 'auto_action', 'action': 'title', 'status': 'skipped', 'detail': blk3['detail']})}\n\n"
            elif assistant_saved:
                blk3 = _aa_block("title", "skipped", {"reason": "conversation already titled"})
                activity_blocks.append(blk3)
                yield f"data: {json.dumps({'type': 'auto_action', 'action': 'title', 'status': 'skipped', 'detail': blk3['detail']})}\n\n"

            # Persist activity blocks to the saved assistant message (so reload shows them)
            if assistant_saved and activity_blocks:
                try:
                    from sqlalchemy import select as _select
                    from backend.database.models import Message as _Msg
                    result = await db.execute(_select(_Msg).where(_Msg.id == saved_msg_id))
                    msg_row = result.scalar_one_or_none()
                    if msg_row is not None:
                        meta = dict(msg_row.extra_metadata or {})
                        existing_blocks = list(meta.get("blocks") or [])
                        # Append only auto_action blocks (avoid duplicating content/thinking)
                        meta["blocks"] = existing_blocks + activity_blocks
                        msg_row.extra_metadata = meta
                        # Also surface thinking/content for backward compat already handled
                        await db.commit()
                        print(f"[AUTO_ACTION] persisted {len(activity_blocks)} blocks to message {saved_msg_id[:8]}")
                except Exception as e:
                    print(f"[AUTO_ACTION] persist failed: {e}")
                    import traceback as _tb3
                    _tb3.print_exc()

            # Yield done event (include real ids so frontend can fix placeholder after regenerate)
            yield f"data: {json.dumps({'type': 'done', 'message_id': saved_msg_id, 'version_group': saved_version_group, 'version': version})}\n\n"
    except asyncio.CancelledError:
        # Request was cancelled by client (Stop button or fallback abort after
        # steer grace window) — flush partial and clear any queued steer so
        # the next turn doesn't see a stale flag.
        print(f"Request {request_id} cancelled by client")
        try:
            await _clear_steer_pending(conversation_id)
        except Exception:
            pass
        await _persist_partial_turn(
            conversation_id, assistant_message, thinking_content,
            message_blocks, model, version, version_group,
            turn_index, placeholder_id, turn_metrics
        )
        raise  # Re-raise to properly propagate cancellation
    except Exception as e:
        # Any uncaught failure (mid-stream connection reset, DB hiccup, provider
        # blow-up between the inner guards) used to kill this generator BEFORE
        # 'done' — leaving clients stuck on the spinner with a never-patched
        # placeholder id. Persist what we have, then ALWAYS terminate the SSE
        # protocol cleanly: an explicit error event followed by done.
        print(f"Error in event generator: {e}")
        import traceback
        traceback.print_exc()
        try:
            await _clear_steer_pending(conversation_id)
        except Exception:
            pass
        saved_id = await _persist_partial_turn(
            conversation_id, assistant_message, thinking_content,
            message_blocks, model, version, version_group,
            turn_index, placeholder_id, turn_metrics
        )
        try:
            yield f"data: {json.dumps({'type': 'error', 'error': str(e)})}\n\n"
        except Exception:
            pass  # Client may have disconnected
        try:
            yield f"data: {json.dumps({'type': 'done', 'message_id': saved_id or placeholder_id, 'version_group': version_group, 'version': version})}\n\n"
        except Exception:
            pass  # Client gone mid-write; nothing more we can do
    finally:
        await _unregister_stream(conversation_id, request_id)

@app.get("/api/stream/{request_id}")
async def stream_response(
    request_id: str,
    conversation_id: str,
    enable_rag: bool = False,
    model: str = None,
    document_ids: str = None,
    provider_id: str = None,
    override_servers: str = None,
    thinking_mode: str = None
):
    """Stream LLM response with real-time tool execution updates."""
    doc_ids = document_ids.split(",") if document_ids else None
    overrides = override_servers.split(",") if override_servers else None
    return StreamingResponse(
        _core_stream_handler(request_id, conversation_id, enable_rag, model, doc_ids,
                             provider_id=provider_id, override_servers=overrides,
                             thinking_mode=thinking_mode),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        }
    )


@app.get("/api/stream/regenerate/{request_id}")
async def stream_regenerate_response(
    request_id: str,
    conversation_id: str,
    model: str = None,
    version: int = 1,
    version_group: str = None,
    anchor_message_id: str = None,
    turn_index: float = None,
    provider_id: str = None,
    override_servers: str = None,
    thinking_mode: str = None
):
    """Stream regenerated LLM response using unified handler.
    Supports versioned regeneration through version/version_group params.
    """
    overrides = override_servers.split(",") if override_servers else None
    return StreamingResponse(
        _core_stream_handler(
            request_id, conversation_id,
            enable_rag=False, model=model,
            version=version, version_group=version_group,
            anchor_message_id=anchor_message_id, turn_index=turn_index,
            provider_id=provider_id, override_servers=overrides,
            thinking_mode=thinking_mode
        ),
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        }
    )


# MCP Registry Endpoints
@app.get("/api/mcp/registry")
async def search_mcp_registry(query: str = "", limit: int = 24):
    """Search installable MCP servers on the Smithery registry. Blank = most used."""
    from tools.mcp_registry import search_mcp_registry, enrich_mcp_servers
    try:
        servers = await search_mcp_registry(query, limit)
        servers = await enrich_mcp_servers(servers)
        return {"servers": servers, "error": None}
    except Exception as e:
        print(f"[MCP REGISTRY] search failed: {e}")
        return {"servers": [], "error": str(e)[:300]}


@app.post("/api/mcp/registry/install")
async def install_mcp_registry_server(request: Request):
    """Install an MCP server from the registry by its qualified name."""
    from tools.mcp_registry import install_mcp_from_registry
    data = await request.json()
    qualified_name = (data.get("qualified_name") or data.get("id") or "").strip()
    if not qualified_name:
        raise HTTPException(status_code=400, detail="Missing server id")
    try:
        result = await install_mcp_from_registry(qualified_name, mcp_manager)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        print(f"[MCP REGISTRY] install failed: {e}")
        raise HTTPException(status_code=500, detail=f"Install failed: {str(e)[:300]}")
    return {"server": result}


# MCP Server Management
@app.get("/api/mcp/servers")
async def list_mcp_servers():
    """List all available MCP servers"""
    # Get runtime server status from MCP manager
    runtime_servers = await mcp_manager.list_servers()

    # Get all servers from database (both enabled and disabled)
    async with get_db() as db:
        db_servers = await get_all_mcp_servers(db)
        db_enabled = {s["name"]: s["enabled"] for s in db_servers}

    # Merge runtime info with enabled status and disabled_tools
    servers_with_status = []
    for server in runtime_servers:
        db_info = next((s for s in db_servers if s["name"] == server["name"]), {})
        servers_with_status.append({
            **server,
            "enabled": db_enabled.get(server["name"], True),
            "disabled_tools": db_info.get("disabled_tools", []),
            "headers": db_info.get("headers", {}),
            "timeout": db_info.get("timeout", 60.0)
        })

    # Also include servers from DB that might not be connected
    for db_server in db_servers:
        if not any(s["name"] == db_server["name"] for s in runtime_servers):
            servers_with_status.append({
                "name": db_server["name"],
                "transport_type": db_server.get("transport_type", "stdio"),
                "command": db_server.get("command"),
                "args": db_server.get("args"),
                "env": db_server.get("env"),
                "url": db_server.get("url"),
                "headers": db_server.get("headers", {}),
                "tool_count": 0,
                "is_connected": False,
                "is_initialized": False,
                "error": None,
                "enabled": db_server.get("enabled", True),
                "disabled_tools": db_server.get("disabled_tools", [])
            })

    return {"servers": servers_with_status}


@app.post("/api/mcp/servers")
async def add_mcp_server(request: Request):
    """Add a new MCP server"""
    data = await request.json()
    name = data.get("name")
    command = data.get("command")
    args = data.get("args", [])
    env = data.get("env", {})
    transport_type = data.get("transport_type", "stdio")
    url = data.get("url")
    timeout = data.get("timeout", 60.0)
    headers = data.get("headers", {}) or {}
    
    # Validate based on transport type
    if transport_type in ("sse", "streamable-http"):
        if not url:
            raise HTTPException(status_code=400, detail="URL is required for SSE/StreamableHTTP transport")
    elif transport_type == "stdio":
        if not command:
            raise HTTPException(status_code=400, detail="Command is required for stdio transport")

    success, error = await mcp_manager.add_server(name, command, args, env, transport_type, url, timeout=timeout, headers=headers)

    if success:
        return {"status": "success", "message": f"Server '{name}' added and connected successfully", "connected": True}
    else:
        return {
            "status": "warning",
            "message": f"Server '{name}' added but connection failed.",
            "connected": False,
            "error": error
        }


@app.delete("/api/mcp/servers/{server_name}")
async def remove_mcp_server(server_name: str):
    """Remove an MCP server"""
    success = await mcp_manager.remove_server(server_name)

    if success:
        return {"status": "success", "message": f"Server '{server_name}' removed"}
    else:
        raise HTTPException(status_code=404, detail="Server not found")


@app.put("/api/mcp/servers/{server_name}")
async def update_mcp_server(server_name: str, request: Request):
    """
    Update an existing MCP server configuration.
    
    This updates the server config and reconnects with new settings.
    """
    data = await request.json()
    
    # Get new configuration
    new_name = data.get("name", server_name)
    command = data.get("command")
    args = data.get("args", [])
    env = data.get("env", {})
    transport_type = data.get("transport_type", "stdio")
    url = data.get("url")
    timeout = data.get("timeout", 60.0)
    headers = data.get("headers", {}) or {}
    
    # Validate based on transport type
    if transport_type in ("sse", "http", "streamable-http"):
        if not url:
            raise HTTPException(status_code=400, detail="URL is required for SSE/HTTP transport")
    elif transport_type == "stdio":
        if not command:
            raise HTTPException(status_code=400, detail="Command is required for stdio transport")
    
    # First remove the old server
    await mcp_manager.remove_server(server_name)
    
    # Add with new configuration (using new name if provided)
    success, error = await mcp_manager.add_server(
        name=new_name,
        command=command,
        args=args,
        env=env,
        transport_type=transport_type,
        url=url,
        timeout=timeout,
        headers=headers
    )

    if success:
        return {"status": "success", "message": f"Server '{new_name}' updated successfully", "connected": True}
    else:
        return {
            "status": "warning",
            "message": f"Server '{new_name}' updated but connection failed.",
            "connected": False,
            "error": error
        }


@app.post("/api/mcp/servers/{server_name}/refresh")
async def refresh_mcp_server_tools(server_name: str):
    """
    Refresh the tool list for a specific MCP server.
    
    Useful when server tools have changed without restarting.
    """
    if server_name not in [s["name"] for s in await mcp_manager.list_servers()]:
        raise HTTPException(status_code=404, detail="Server not found")
    
    success = await mcp_manager.refresh_tools(server_name)
    
    if success:
        return {"status": "success", "message": f"Tools refreshed for server '{server_name}'"}
    else:
        raise HTTPException(status_code=500, detail="Failed to refresh tools")


@app.post("/api/mcp/servers/{server_name}/reconnect")
async def reconnect_mcp_server(server_name: str):
    """
    Reconnect to an MCP server.

    Useful when connection was lost or server was restarted.
    """
    if server_name not in [s["name"] for s in await mcp_manager.list_servers()]:
        raise HTTPException(status_code=404, detail="Server not found")

    success = await mcp_manager.reconnect_server(server_name)

    if success:
        return {"status": "success", "message": f"Reconnected to server '{server_name}'"}
    else:
        raise HTTPException(status_code=500, detail="Failed to reconnect to server")


@app.post("/api/mcp/servers/{server_name}/toggle")
async def toggle_mcp_server_endpoint(server_name: str, request: Request):
    """
    Enable or disable an MCP server.
    When disabling, disconnects the server from the runtime manager.
    When enabling, reconnects the server with stored config.
    """
    from database.crud import toggle_mcp_server as db_toggle_mcp_server
    from database.crud import get_all_mcp_servers

    data = await request.json()
    enabled = data.get("enabled", True)

    async with get_db() as db:
        await db_toggle_mcp_server(db, server_name, enabled)

    if enabled:
        # Reconnect: fetch config from DB and connect
        async with get_db() as db:
            all_servers = await get_all_mcp_servers(db)
        config_data = next((s for s in all_servers if s["name"] == server_name), None)
        if config_data:
            config = MCPServerConfig(
                name=config_data["name"],
                transport_type=config_data.get("transport_type", "stdio"),
                command=config_data.get("command"),
                args=config_data.get("args", []),
                env=config_data.get("env", {}),
                url=config_data.get("url"),
                headers=config_data.get("headers", {}),
                timeout=config_data.get("timeout", 60.0)
            )
            await mcp_manager._connect_server(config)
    else:
        # Disconnect: remove from manager but keep in DB
        if server_name in mcp_manager.servers:
            instance = mcp_manager.servers[server_name]
            if instance.client and instance.is_connected:
                try:
                    await instance.client.close()
                except Exception:
                    pass
            del mcp_manager.servers[server_name]

    return {"status": "success", "message": f"Server '{server_name}' {'enabled' if enabled else 'disabled'}"}


@app.put("/api/mcp/servers/{server_name}/tools/toggle")
async def toggle_mcp_tool_endpoint(server_name: str, request: Request):
    """
    Enable or disable a specific tool on an MCP server.
    Disabled tools are excluded from the LLM prompt, saving context tokens.
    """
    data = await request.json()
    tool_name = data.get("tool_name")
    disabled = data.get("disabled", True)
    
    if not tool_name:
        raise HTTPException(status_code=400, detail="tool_name is required")
    
    # Update in database
    async with get_db() as db:
        await update_mcp_server_disabled_tools(db, server_name, tool_name, disabled)
    
    # Update runtime config so filtering takes effect immediately
    if server_name in mcp_manager.servers:
        instance = mcp_manager.servers[server_name]
        current = set(instance.config.disabled_tools or [])
        if disabled:
            current.add(tool_name)
        else:
            current.discard(tool_name)
        instance.config.disabled_tools = list(current)
    
    return {
        "status": "success",
        "message": f"Tool '{tool_name}' {'disabled' if disabled else 'enabled'} on '{server_name}'"
    }


# Conversation Management
@app.delete("/api/conversations/{conversation_id}")
async def delete_conversation(conversation_id: str):
    """Delete a conversation"""
    async with get_db() as db:
        await db_delete_conversation(db, conversation_id)
        return {"status": "success", "message": "Conversation deleted"}


@app.put("/api/conversations/{conversation_id}")
async def update_conversation(conversation_id: str, request: Request):
    """Update a conversation title"""
    data = await request.json()
    title = data.get("title", "")
    
    if not title.strip():
        raise HTTPException(status_code=400, detail="Title cannot be empty")
    
    async with get_db() as db:
        await update_conversation_title(db, conversation_id, title)
        return {"status": "success", "message": "Conversation updated"}


@app.put("/api/messages/{message_id}")
async def edit_message(message_id: str, request: Request):
    """Edit a message's content"""
    data = await request.json()
    content = data.get("content", "")
    
    if not content.strip():
        raise HTTPException(status_code=400, detail="Content cannot be empty")
    
    async with get_db() as db:
        message = await update_message(db, message_id, content)
        if not message:
            raise HTTPException(status_code=404, detail="Message not found")
        return {"message": message}


@app.get("/api/messages/{message_id}/versions")
async def get_message_versions_endpoint(message_id: str):
    """Get all versions of a message by message_id.
    Returns empty list if the message has no version_group.
    """
    async with get_db() as db:
        versions = await get_message_versions(db, message_id)
        return {"versions": versions}


@app.get("/api/versions/{version_group}")
async def get_versions_by_group(version_group: str):
    """Get all versions in a version_group.
    Uses the version_group directly (no message_id lookup needed).
    """
    from sqlalchemy import select
    from database.models import Message
    async with get_db() as db:
        result = await db.execute(
            select(Message)
            .where(Message.version_group == version_group)
            .order_by(Message.version)
        )
        versions_obj = result.scalars().all()
        versions = [
            {
                "id": v.id,
                "role": v.role,
                "content": v.content,
                "tool_calls": v.tool_calls,
                "thinking": v.thinking,
                "metadata": v.extra_metadata,
                "blocks": v.extra_metadata.get('blocks') if v.extra_metadata else None,
                "version": v.version,
                "version_group": v.version_group,
                "created_at": v.created_at.isoformat(),
            }
            for v in versions_obj
        ]
        return {"versions": versions}


@app.delete("/api/messages/{message_id}")
async def delete_message_endpoint(message_id: str, version: Optional[int] = None):
    """Delete a message, or (with ?version=N) one specific version of a
    versioned response — that row's id is resolved from the message's group."""
    async with get_db() as db:
        success = await db_delete_message(db, message_id, version=version)
        if not success:
            raise HTTPException(status_code=404, detail="Message not found")
        return {"status": "success", "message": "Message deleted"}


@app.post("/api/conversations/{conversation_id}/regenerate")
async def regenerate_last_response(conversation_id: str, request: Request):
    """Regenerate an assistant response with versioning.
    Instead of deleting old messages, creates a new version of the response.
    """
    import uuid as uuid_lib
    data = await request.json()
    message_id = data.get("message_id")
    
    async with get_db() as db:
        # Get the message to regenerate
        if message_id:
            message = await get_message(db, message_id)
            if not message:
                raise HTTPException(status_code=404, detail="Message not found")
            if message.get("role") != "assistant":
                raise HTTPException(status_code=400, detail="Can only regenerate assistant messages")
        else:
            # Get last assistant message if no message_id provided
            messages = await get_conversation_messages(db, conversation_id, only_latest_versions=False)
            # Find last assistant message
            assistant_messages = [m for m in messages if m.get("role") == "assistant"]
            if not assistant_messages:
                raise HTTPException(status_code=400, detail="No assistant message to regenerate")
            message = assistant_messages[-1]
        
        # Find the user message that preceded this assistant message
        # Use all messages (not just latest versions) to find the right position
        messages = await get_conversation_messages(db, conversation_id, only_latest_versions=False)
        
        # Find the index of the current message (last version in its group if it has a group)
        msg_target_id = message.get("id")
        msg_index = -1
        for i, m in enumerate(messages):
            if m.get("id") == msg_target_id:
                msg_index = i
                break
        
        if msg_index <= 0:
            raise HTTPException(status_code=400, detail="Could not find preceding user message")
        
        # Walk backwards from msg_index to find the actual preceding user message,
        # skipping any old versions of the same assistant response (same version_group)
        target_version_group = message.get("version_group")
        user_message = None
        for i in range(msg_index - 1, -1, -1):
            m = messages[i]
            if m.get("role") == "user":
                user_message = m
                break
            # If this is an old version of the same response, skip it
            if m.get("version_group") == target_version_group:
                continue
        
        if not user_message:
            raise HTTPException(status_code=400, detail="Could not find preceding user message")
        
        # Determine version_group, next version number, and the timeline slot.
        from sqlalchemy import select, func
        from database.models import Message as MsgModel

        # ORM row of the response being replaced — source of the turn slot and
        # the object stamped with a fresh version_group on first regeneration.
        sup_res = await db.execute(select(MsgModel).where(MsgModel.id == msg_target_id))
        superseded = sup_res.scalar_one_or_none()

        current_version_group = message.get("version_group")

        if current_version_group:
            version_group = current_version_group
            # Derive from MAX(version) so concurrent/stale regenerations of any
            # group member can never mint duplicate (group, version) pairs.
            max_v = (await db.execute(
                select(func.max(MsgModel.version)).where(MsgModel.version_group == version_group)
            )).scalar()
            new_version = (max_v or message.get("version", 1)) + 1
        else:
            # First regenerate — create the version group on the original.
            version_group = str(uuid_lib.uuid4())
            new_version = message.get("version", 1) + 1
            if superseded:
                superseded.version_group = version_group

        # All versions occupy the superseded response's timeline slot so the
        # regenerated answer renders (and replays history) at the same position.
        turn_index = superseded.turn_index if superseded else None
        if turn_index is None:
            turn_index = (user_message.get("turn_index") or 0.0) + 0.5

        await db.commit()

        # Create new request ID
        request_id = str(uuid.uuid4())

        return {
            "request_id": request_id,
            "status": "processing",
            "conversation_id": conversation_id,
            "version": new_version,
            "version_group": version_group,
            "anchor_user_message_id": user_message["id"],
            "turn_index": turn_index
        }


# Chat File Upload
@app.post("/api/upload/chat-file")
async def upload_chat_file(file: UploadFile = File(...)):
    """Upload a file for use in a chat message (images, documents, etc.)"""
    if not file.filename:
        raise HTTPException(status_code=400, detail="No file provided")

    # Read file content
    content = await file.read()
    file_size = len(content)

    try:
        max_size = int(settings_manager.get_settings().get("max_upload_size", MAX_UPLOAD_SIZE))
    except Exception:
        max_size = MAX_UPLOAD_SIZE
    if file_size > max_size:
        raise HTTPException(status_code=400, detail=f"File too large ({file_size} bytes). Max size: {max_size} bytes ({max_size // (1024*1024)} MB)")

    # Detect content type from file or use provided
    import mimetypes
    content_type = file.content_type or mimetypes.guess_type(file.filename)[0] or "application/octet-stream"
    
    # Generate unique filename preserving extension
    file_ext = os.path.splitext(file.filename)[1].lower()
    unique_name = f"{uuid.uuid4()}{file_ext}"
    file_path = os.path.join(UPLOAD_DIR, unique_name)
    
    # Save file
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    with open(file_path, "wb") as f:
        f.write(content)
    
    return {
        "url": f"/uploads/{unique_name}",
        "filename": file.filename,
        "type": content_type,
        "size": file_size
    }


# Document Management
@app.post("/api/documents/upload")
async def upload_document(background_tasks: BackgroundTasks, file: UploadFile = File(...)):
    """Upload a document to the knowledgebase and process it for RAG"""
    raw_name = (file.filename or "").strip()
    if not raw_name:
        raise HTTPException(status_code=400, detail="No file provided")

    # Read content first — single source of truth for size (avoids SpooledTemporaryFile seek quirks)
    content = await file.read()
    file_size = len(content)

    # Dynamic limit so changes via Settings → Save take effect without restart
    try:
        max_size = int(settings_manager.get_settings().get("max_upload_size", MAX_UPLOAD_SIZE))
    except Exception:
        max_size = MAX_UPLOAD_SIZE

    if file_size == 0:
        raise HTTPException(status_code=400, detail="File is empty")
    if file_size > max_size:
        raise HTTPException(
            status_code=400,
            detail=f"File too large ({file_size} bytes). Max size: {max_size} bytes ({max_size // (1024*1024)} MB). Increase it in Settings or set MAX_UPLOAD_SIZE env var."
        )

    # Robust extension detection: strip, lower, fallback to content-type
    file_ext = os.path.splitext(raw_name)[1].lower().strip()
    content_type = (file.content_type or "").lower().strip()
    # Fallback when browser omits extension (common on mobile) or uses weird casing/spaces
    if file_ext not in [".txt", ".md", ".pdf", ".docx", ".json", ".yaml", ".yml"]:
        if "pdf" in content_type:
            file_ext = ".pdf"
        elif "officedocument" in content_type or "msword" in content_type or raw_name.lower().endswith(".docx"):
            file_ext = ".docx"
        elif "json" in content_type:
            file_ext = ".json"
        elif "yaml" in content_type or "yml" in content_type:
            file_ext = ".yaml"
        elif content_type.startswith("text/"):
            file_ext = ".txt"

    # Determine file type BEFORE writing to disk
    file_type = "unknown"
    if file_ext in [".txt", ".md"]:
        file_type = "text"
    elif file_ext == ".pdf":
        file_type = "pdf"
    elif file_ext == ".docx":
        file_type = "document"
    elif file_ext in [".json", ".yaml", ".yml"]:
        file_type = "data"

    if file_type == "unknown":
        print(f"[UPLOAD] rejected {raw_name!r} ext={file_ext!r} content_type={content_type!r} size={file_size}")
        raise HTTPException(status_code=400, detail=f"Unsupported file type '{file_ext or '(none)'}'. Supported: txt, md, pdf, docx, json, yaml, yml")

    print(f"[UPLOAD] {raw_name!r} -> {file_ext} ({file_type}) {file_size} bytes, content_type={content_type!r}")

    # Create upload directory if it doesn't exist
    os.makedirs(UPLOAD_DIR, exist_ok=True)

    # Generate unique filename preserving (now-normalized) extension
    unique_filename = f"{uuid.uuid4()}{file_ext}"
    file_path = os.path.join(UPLOAD_DIR, unique_filename)

    # Save the file
    with open(file_path, "wb") as f:
        f.write(content)

    # Create document record
    async with get_db() as db:
        document = await create_document(
            db,
            filename=raw_name,
            filepath=file_path,
            file_type=file_type,
            size_bytes=file_size,
            metadata={"original_filename": raw_name}
        )

        # Mark as processing
        await update_document_status(db, document["id"], "processing", {})

        document_id = document["id"]

    # Process document for RAG in background
    background_tasks.add_task(
        process_document_background,
        document_id,
        file_path,
        file_type
    )

    return {
        "status": "processing",
        "document": {
            "id": document_id,
            "filename": raw_name,
            "file_type": file_type,
            "size_bytes": file_size,
            "status": "processing"
        }
    }


async def process_document_background(document_id: str, file_path: str, file_type: str):
    """Background task to process document for RAG"""
    async with get_db() as db:
        try:
            # Process document for RAG
            result = await tool_executor.process_document_for_rag(
                document_id=document_id,
                filepath=file_path,
                file_type=file_type
            )
            
            if result.get("success"):
                await update_document_status(
                    db,
                    document_id,
                    "completed",
                    {"chunks": result.get("chunk_count", 0)}
                )
            else:
                await update_document_status(
                    db,
                    document_id,
                    "failed",
                    {"error": result.get("error", "Unknown error")}
                )
        except Exception as e:
            await update_document_status(
                db,
                document_id,
                "failed",
                {"error": str(e)}
            )


@app.get("/api/documents")
async def list_documents():
    """List all documents in knowledgebase"""
    async with get_db() as db:
        documents = await get_documents(db)
        return {"documents": documents}


@app.get("/api/documents/{document_id}")
async def get_document_detail(document_id: str):
    """Get document details"""
    async with get_db() as db:
        document = await get_document(db, document_id)
        if not document:
            raise HTTPException(status_code=404, detail="Document not found")
        return {"document": document}


@app.delete("/api/documents/{document_id}")
async def delete_document_endpoint(document_id: str):
    """Delete a document from knowledgebase"""
    async with get_db() as db:
        # Get document to find filepath
        document = await get_document(db, document_id)
        if not document:
            raise HTTPException(status_code=404, detail="Document not found")
        
        # Delete from RAG index
        await tool_executor.delete_document_from_rag(document_id)
        
        # Delete file from disk
        try:
            if os.path.exists(document["filepath"]):
                os.remove(document["filepath"])
        except Exception as e:
            print(f"Error deleting file: {e}")
        
        # Delete from database
        success = await db_delete_document(db, document_id)
        if not success:
            raise HTTPException(status_code=404, detail="Document not found")
        return {"status": "success", "message": "Document deleted"}


@app.get("/api/tools")
async def list_custom_tool_catalogue():
    """The authoritative custom-tool catalogue for the agent capability picker.

    Served from the same definitions the executor dispatches, so the picker can
    never drift from what an agent is actually allowed to call.
    """
    return {"tools": custom_tool_catalogue()}


@app.get("/api/mcp/tools")
async def list_available_tools():
    """List all tools from all MCP servers and custom tools"""
    # Get MCP tools
    mcp_tools = []
    if mcp_manager:
        mcp_tools = await mcp_manager.list_all_tools()
    
    # Get custom tool definitions
    custom_tools = tool_executor.get_tool_definitions()
    
    return {
        "tools": mcp_tools,
        "custom_tools": custom_tools
    }


@app.get("/api/models")
async def list_available_models():
    """List models across all enabled LLM providers (cached at connect/refresh).

    Each model carries `context_window` when it is known (see
    _backfill_context_windows at startup). Absent means unknown — the UI shows a
    bare token count instead of a fake utilisation percentage.
    """
    from database.provider_crud import list_providers
    async with get_db() as db:
        providers = await list_providers(db)
    models = []
    provider_list = []
    for p in providers:
        if not p.get("enabled"):
            continue
        provider_list.append({"id": p["id"], "name": p["name"], "is_default": p.get("is_default", 0)})
        for m in (p.get("models") or []):
            models.append({
                "id": m.get("id"),
                "name": m.get("name") or m.get("id"),
                "owned_by": m.get("owned_by") or p.get("name"),
                "provider_id": p.get("id"),
                "provider_name": p.get("name"),
                "context_window": m.get("context_window"),
            })
    return {"models": models, "providers": provider_list}


@app.post("/api/rag/query")
async def rag_query_endpoint(request: Request):
    """
    Direct RAG query endpoint for searching documents.
    
    This can be used for explicit document queries without LLM tool calling.
    """
    data = await request.json()
    query = data.get("query", "")
    document_ids = data.get("document_ids")
    section = data.get("section")
    top_k = min(data.get("top_k", 10), 50)
    
    if not query:
        raise HTTPException(status_code=400, detail="Query is required")
    
    result = await tool_executor.rag_service.query(
        query=query,
        document_ids=document_ids,
        top_k=top_k,
        section=section
    )
    
    return result


# Agent Management
def _agent_to_dict(agent) -> Dict:
    """Serialize an Agent model to a dictionary."""
    return {
        "id": agent.id,
        "name": agent.name,
        "description": agent.description,
        "model": agent.model,
        "provider_id": getattr(agent, "provider_id", None),
        "temperature": agent.temperature,
        "top_k": agent.top_k,
        "max_tokens": agent.max_tokens,
        "system_prompt": agent.system_prompt,
        "enabled_tools": agent.enabled_tools,
        "enabled_mcp_servers": agent.enabled_mcp_servers,
        "enabled_skills": agent.enabled_skills or [],
        "enable_rag": bool(agent.enable_rag),
        "rag_similarity_threshold": agent.rag_similarity_threshold,
        "enable_web_search": bool(agent.enable_web_search),
        "conversation_starters": agent.conversation_starters,
        "created_at": agent.created_at.isoformat(),
        "updated_at": agent.updated_at.isoformat(),
        "is_active": bool(agent.is_active)
    }


@app.get("/api/agents")
async def list_agents():
    """List all agents"""
    async with get_db() as db:
        agents = await get_all_agents(db)
        return {"agents": [_agent_to_dict(a) for a in agents]}


@app.get("/api/agents/{agent_id}")
async def get_agent_detail(agent_id: int):
    """Get agent details"""
    async with get_db() as db:
        agent = await get_agent(db, agent_id)
        if not agent:
            raise HTTPException(status_code=404, detail="Agent not found")
        return {"agent": _agent_to_dict(agent)}


@app.post("/api/agents")
async def create_agent_endpoint(request: Request):
    """Create a new agent"""
    data = await request.json()
    
    agent_data = {
        "name": data.get("name"),
        "description": data.get("description", ""),
        "model": data.get("model", "qwen3-4b"),
        "provider_id": data.get("provider_id"),
        "temperature": data.get("temperature", 0.7),
        "top_k": data.get("top_k", 40),
        "max_tokens": data.get("max_tokens", 16048),
        "system_prompt": data.get("system_prompt", ""),
        "enabled_tools": data.get("enabled_tools", []),
        "enabled_mcp_servers": data.get("enabled_mcp_servers", []),
        "enabled_skills": data.get("enabled_skills", []),
        "enable_rag": 1 if data.get("enable_rag", False) else 0,
        "rag_similarity_threshold": data.get("rag_similarity_threshold", 0.4),
        "enable_web_search": 1 if data.get("enable_web_search", False) else 0,
        "conversation_starters": data.get("conversation_starters", [])
    }
    
    # Validate required fields
    if not agent_data["name"]:
        raise HTTPException(status_code=400, detail="Agent name is required")
    
    async with get_db() as db:
        # Check if name already exists
        existing = await get_agent_by_name(db, agent_data["name"])
        if existing:
            raise HTTPException(status_code=400, detail="Agent with this name already exists")
        
        agent = await create_agent(db, agent_data)
        return {"agent": _agent_to_dict(agent)}


@app.put("/api/agents/{agent_id}")
async def update_agent_endpoint(agent_id: int, request: Request):
    """Update an agent"""
    data = await request.json()
    
    update_data = {}
    if "name" in data:
        update_data["name"] = data["name"]
    if "description" in data:
        update_data["description"] = data["description"]
    if "model" in data:
        update_data["model"] = data["model"]
    if "provider_id" in data:
        update_data["provider_id"] = data["provider_id"] or None
    if "temperature" in data:
        update_data["temperature"] = data["temperature"]
    if "top_k" in data:
        update_data["top_k"] = data["top_k"]
    if "max_tokens" in data:
        update_data["max_tokens"] = data["max_tokens"]
    if "system_prompt" in data:
        update_data["system_prompt"] = data["system_prompt"]
    if "enabled_tools" in data:
        update_data["enabled_tools"] = data["enabled_tools"]
    if "enabled_mcp_servers" in data:
        update_data["enabled_mcp_servers"] = data["enabled_mcp_servers"]
    if "enabled_skills" in data:
        update_data["enabled_skills"] = data["enabled_skills"]
    if "enable_rag" in data:
        update_data["enable_rag"] = 1 if data["enable_rag"] else 0
    if "rag_similarity_threshold" in data:
        update_data["rag_similarity_threshold"] = data["rag_similarity_threshold"]
    if "enable_web_search" in data:
        update_data["enable_web_search"] = 1 if data["enable_web_search"] else 0
    if "conversation_starters" in data:
        update_data["conversation_starters"] = data["conversation_starters"]
    
    async with get_db() as db:
        agent = await update_agent(db, agent_id, update_data)
        if not agent:
            raise HTTPException(status_code=404, detail="Agent not found")
        return {"agent": _agent_to_dict(agent)}


@app.delete("/api/agents/{agent_id}")
async def delete_agent_endpoint(agent_id: int):
    """Delete an agent (soft delete)"""
    async with get_db() as db:
        success = await delete_agent(db, agent_id)
        if not success:
            raise HTTPException(status_code=404, detail="Agent not found")
        return {"status": "success", "message": "Agent deleted"}


# TTS Endpoints
@app.post("/api/tts/generate")
async def generate_tts(request: Request):
    """
    Generate speech audio from text using TTS.
    
    Returns audio file URL that can be played in the browser.
    """
    try:
        data = await request.json()
        text = data.get("text", "")
        voice = data.get("voice")
        
        if not text.strip():
            raise HTTPException(status_code=400, detail="Text is required")

        # Watch for client disconnect (pause/stop clicked): when it happens, set
        # the stop flag so long Kokoro generation bails between segments.
        stop_flag = threading.Event()
        watcher = asyncio.create_task(_watch_disconnect(request, stop_flag))
        try:
            result = await tool_executor.tts_service.generate_speech(
                text=text,
                voice=voice,
                should_stop=stop_flag.is_set
            )
        finally:
            watcher.cancel()
        
        if not result.get("success"):
            raise HTTPException(status_code=500, detail=result.get("error", "TTS generation failed"))
        
        return result
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"TTS error: {str(e)}")


@app.post("/api/tts/generate/stream")
async def generate_tts_stream(request: Request):
    """Stream TTS audio as it is generated.

    Response is NDJSON (one JSON object per line, media_type application/x-ndjson):
      {"seg": "tts_xxx.wav", "url": "/api/audio/tts_xxx.wav"}   — one segment ready
      {"done": true}                                             — all segments sent
      {"error": "..."}                                          — failure

    The client plays segments as they arrive (first sentence ≈ first audio).
    Cancelling the fetch marks the request disconnected; the server stops
    generation between sentences.
    """
    from tools.tts_service import HAS_EDGE_TTS, _check_kokoro_available
    try:
        data = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid JSON body")
    text = data.get("text", "")
    voice = data.get("voice")
    if not text.strip():
        raise HTTPException(status_code=400, detail="Text is required")

    svc = tool_executor.tts_service
    engine = svc.config.engine
    if engine == "kokoro" and not _check_kokoro_available():
        raise HTTPException(status_code=500, detail="Kokoro engine is not available")
    if engine == "edge-tts" and not HAS_EDGE_TTS:
        raise HTTPException(status_code=500, detail="Edge TTS engine is not available")

    stop_flag = threading.Event()
    watcher = asyncio.create_task(_watch_disconnect(request, stop_flag))

    async def event_stream():
        try:
            async for filename, url in svc.stream_speech(
                text=text, voice=voice, should_stop=stop_flag.is_set
            ):
                yield json.dumps({"seg": filename, "url": url}) + "\n"
            yield json.dumps({"done": True}) + "\n"
        except Exception as e:
            import traceback
            traceback.print_exc()
            yield json.dumps({"error": str(e)}) + "\n"
        finally:
            watcher.cancel()

    return StreamingResponse(
        event_stream(),
        media_type="application/x-ndjson",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


async def _watch_disconnect(request: Request, stop_flag):
    """Poll request.is_disconnected() from the event loop and set stop_flag."""
    while not stop_flag.is_set():
        if await request.is_disconnected():
            stop_flag.set()
            return
        await asyncio.sleep(0.2)


@app.get("/api/tts/voices")
async def list_tts_voices():
    """List available TTS voices"""
    return tool_executor.tts_service.list_available_voices()


@app.get("/api/tts/status")
async def get_tts_status():
    """Check if TTS is available"""
    from tools.tts_service import HAS_EDGE_TTS, _check_kokoro_available
    kokoro_available = _check_kokoro_available()
    return {
        "available": HAS_EDGE_TTS or kokoro_available,
        "edge_tts": HAS_EDGE_TTS,
        "kokoro": kokoro_available,
        "engine": tool_executor.tts_service.config.engine
    }


# Memory Endpoints (agent platform Phase 2)
@app.get("/api/memory")
async def list_memory(scope: str = None, limit: int = 200):
    """List persistent memory entries (optional scope filter)."""
    from database.memory_crud import list_memory_entries
    async with get_db() as db:
        entries = await list_memory_entries(db, scope=scope, limit=limit)
    return {"entries": entries}


@app.post("/api/memory")
async def add_memory(request: Request):
    """Create a memory entry."""
    from database.memory_crud import create_memory_entry
    data = await request.json()
    content = (data.get("content") or "").strip()
    if not content:
        raise HTTPException(status_code=400, detail="Content required")
    scope = data.get("scope") or "global"
    async with get_db() as db:
        entry = await create_memory_entry(db, content, scope=scope,
                                          tags=data.get("tags") or [], source="manual")
    return {"entry": entry}


@app.patch("/api/memory/{entry_id}")
async def edit_memory(entry_id: str, request: Request):
    """Update a memory entry's content."""
    from database.memory_crud import update_memory_entry
    data = await request.json()
    async with get_db() as db:
        entry = await update_memory_entry(db, entry_id, content=data.get("content"))
    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")
    return {"entry": entry}


@app.delete("/api/memory/{entry_id}")
async def remove_memory(entry_id: str):
    """Delete a memory entry."""
    from database.memory_crud import delete_memory_entry
    async with get_db() as db:
        deleted = await delete_memory_entry(db, entry_id)
    if not deleted:
        raise HTTPException(status_code=404, detail="Entry not found")
    return {"status": "ok"}


# Skill Registry Endpoints (Phase 4.5)
@app.get("/api/skills/registry")
async def search_skill_registry(query: str = "", limit: int = 25):
    """Search installable skills on the skills.sh registry. Blank query = most popular."""
    from tools.skill_registry import search_registry, popular_registry, enrich_registry
    try:
        q = (query or "").strip()
        skills = await (popular_registry(limit) if len(q) < 2 else search_registry(q, limit))
        skills = await enrich_registry(skills)
        return {"skills": skills, "error": None}
    except Exception as e:
        print(f"[REGISTRY] search failed: {e}")
        return {"skills": [], "error": str(e)[:300]}


@app.post("/api/skills/install")
async def install_registry_skill(request: Request):
    """Install a registry skill (skills.sh id, e.g. 'owner/repo/path') locally."""
    from tools.skill_registry import install_registry_skill
    data = await request.json()
    skill_id = (data.get("id") or data.get("skill_id") or "").strip()
    if not skill_id:
        raise HTTPException(status_code=400, detail="Missing skill id")
    try:
        result = await install_registry_skill(skill_id)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        print(f"[REGISTRY] install failed: {e}")
        raise HTTPException(status_code=500, detail=f"Install failed: {str(e)[:300]}")
    return {"skill": result}


# Skills Endpoints (agent platform Phase 3/4)
@app.get("/api/skills")
async def list_skills_api(include_drafts: bool = False):
    """List installed skills (optionally including _drafts)."""
    from tools.skills_tool import list_skills
    return {"skills": list_skills(include_drafts=include_drafts)}


@app.get("/api/skills/{name}")
async def get_skill_api(name: str):
    """Get a skill's SKILL.md content + file manifest."""
    from tools.skills_tool import get_skill
    skill = get_skill(name)
    if not skill:
        raise HTTPException(status_code=404, detail="Skill not found")
    return {"skill": skill}


@app.post("/api/skills")
async def create_skill_api(request: Request):
    """Create a skill."""
    from tools.skills_tool import write_skill
    data = await request.json()
    name = (data.get("name") or "").strip()
    description = (data.get("description") or "").strip()
    instructions = (data.get("instructions") or "").strip()
    if not name or not instructions:
        raise HTTPException(status_code=400, detail="name and instructions are required")
    skill = write_skill(name, description or name, instructions)
    return {"skill": skill}


@app.put("/api/skills/{name}")
async def update_skill_api(name: str, request: Request):
    """Update a skill's description/instructions."""
    from tools.skills_tool import write_skill
    data = await request.json()
    description = (data.get("description") or "").strip()
    instructions = (data.get("instructions") or "").strip()
    if not instructions:
        raise HTTPException(status_code=400, detail="instructions are required")
    skill = write_skill(name, description or name, instructions)
    return {"skill": skill}


@app.delete("/api/skills/{name}")
async def delete_skill_api(name: str):
    """Delete a skill."""
    from tools.skills_tool import delete_skill
    if not delete_skill(name):
        raise HTTPException(status_code=404, detail="Skill not found")
    return {"status": "ok"}


@app.post("/api/skills/drafts/{name}/accept")
async def accept_skill_draft_api(name: str):
    """Accept a self-improvement draft: move it into the live skills dir."""
    from tools.skills_tool import accept_draft
    skill = accept_draft(name)
    if not skill:
        raise HTTPException(status_code=404, detail="Draft not found")
    return {"skill": skill}


@app.delete("/api/skills/drafts/{name}")
async def reject_skill_draft_api(name: str):
    """Reject a self-improvement draft: delete it."""
    from tools.skills_tool import delete_skill
    if not delete_skill(name, draft=True):
        raise HTTPException(status_code=404, detail="Draft not found")
    return {"status": "ok"}


# LLM Provider Endpoints (multi-provider support)
@app.get("/api/providers")
async def list_llm_providers():
    """List LLM providers (with cached models)."""
    from database.provider_crud import list_providers
    async with get_db() as db:
        providers = await list_providers(db)
    return {"providers": providers}


@app.post("/api/providers")
async def add_llm_provider(request: Request):
    """Add a provider and auto-fetch its models from /v1/models."""
    from database.provider_crud import create_provider, get_provider_by_name
    from tools.provider_service import fetch_models
    data = await request.json()
    name = (data.get("name") or "").strip()
    base_url = (data.get("base_url") or "").strip().rstrip("/")
    api_key = (data.get("api_key") or "").strip() or None
    if not name or not base_url:
        raise HTTPException(status_code=400, detail="name and base_url are required")
    async with get_db() as db:
        existing = await get_provider_by_name(db, name)
        if existing:
            raise HTTPException(status_code=409, detail=f"Provider '{name}' already exists")
        # First provider becomes the default.
        from database.provider_crud import list_providers as _lp
        existing_count = len(await _lp(db))
        try:
            models = await fetch_models(base_url, api_key)
            error = None
        except Exception as e:
            models = []
            error = str(e)[:300]
        provider = await create_provider(
            db, name, base_url, api_key=api_key, models=models,
            is_default=1 if existing_count == 0 else 0
        )
        await db.commit()
    return {"provider": provider, "models_fetched": len(models), "error": error}


@app.post("/api/providers/{provider_id}/refresh")
async def refresh_llm_provider_models(provider_id: str):
    """Re-fetch models from a provider."""
    from database.provider_crud import get_provider, update_provider
    from tools.provider_service import fetch_models
    async with get_db() as db:
        provider = await get_provider(db, provider_id, include_api_key=True)
        if not provider:
            raise HTTPException(status_code=404, detail="Provider not found")
        try:
            models = await fetch_models(provider["base_url"], provider.get("api_key"))
            error = None
        except Exception as e:
            models = provider.get("models") or []
            error = str(e)[:300]
        updated = await update_provider(db, provider_id, models=models)
        await db.commit()
    return {"provider": updated, "models_fetched": len(models), "error": error}


@app.post("/api/providers/{provider_id}/default")
async def set_default_llm_provider(provider_id: str):
    """Make a provider the default for conversations/agents without one."""
    from database.provider_crud import set_default_provider
    async with get_db() as db:
        provider = await set_default_provider(db, provider_id)
        if not provider:
            raise HTTPException(status_code=404, detail="Provider not found")
        await db.commit()
    return {"provider": provider}


@app.put("/api/providers/{provider_id}")
async def update_llm_provider(provider_id: str, request: Request):
    """Update provider details and re-fetch models."""
    from database.provider_crud import get_provider, update_provider
    from tools.provider_service import fetch_models
    data = await request.json()
    async with get_db() as db:
        provider = await get_provider(db, provider_id, include_api_key=True)
        if not provider:
            raise HTTPException(status_code=404, detail="Provider not found")
        fields = {}
        if "name" in data and str(data.get("name") or "").strip():
            fields["name"] = str(data["name"]).strip()
        if "base_url" in data and str(data.get("base_url") or "").strip():
            fields["base_url"] = str(data["base_url"]).strip().rstrip("/")
        if "api_key" in data:
            fields["api_key"] = str(data["api_key"]).strip() or None
        if "enabled" in data:
            fields["enabled"] = 1 if data["enabled"] else 0
        if fields.get("base_url") and fields["base_url"] != provider["base_url"]:
            try:
                fields["models"] = await fetch_models(fields["base_url"], fields.get("api_key", provider.get("api_key")))
            except Exception as e:
                fields["models"] = []
                fields["fetch_error"] = str(e)[:300]
        updated = await update_provider(db, provider_id, **fields)
        await db.commit()
    return {"provider": updated}


@app.delete("/api/providers/{provider_id}")
async def delete_llm_provider(provider_id: str):
    """Delete a provider."""
    from database.provider_crud import delete_provider, get_default_provider, set_default_provider, list_providers
    async with get_db() as db:
        provider = await delete_provider(db, provider_id)
        if not provider:
            raise HTTPException(status_code=404, detail="Provider not found")
        # If we deleted the default, promote the first remaining provider.
        remaining = await list_providers(db)
        if remaining and not any(p.get("is_default") for p in remaining):
            await set_default_provider(db, remaining[0]["id"])
        await db.commit()
    return {"status": "ok"}


# Terminal Endpoints
@app.get("/api/terminal/blocked-patterns")
async def get_terminal_blocked_patterns():
    """List the hard-coded dangerous command patterns (read-only, cannot be disabled)."""
    from tools.terminal_tool import HARD_BLOCKED_PATTERNS
    return {"patterns": HARD_BLOCKED_PATTERNS}


# Jobs Endpoints (agent platform Phase 5)
@app.get("/api/jobs")
async def list_jobs_api(limit: int = 50):
    """List recent job runs."""
    from database.job_crud import list_job_runs
    async with get_db() as db:
        runs = await list_job_runs(db, limit=limit)
    return {"runs": runs}


async def _pick_jobs_model() -> Optional[str]:
    """jobs_model setting, else the default provider's first loaded model."""
    from settings import settings_manager
    m = (settings_manager.get_settings().get("jobs_model") or "").strip()
    if m:
        return m
    try:
        from database.provider_crud import get_default_provider
        async with get_db() as db:
            provider = await get_default_provider(db)
        if not provider:
            return None
        cached = provider.get("models") or []
        base = provider.get("base_url")
        import aiohttp
        async with aiohttp.ClientSession(timeout=aiohttp.ClientTimeout(total=10)) as session:
            async with session.get(f"{base}/v1/models") as response:
                if response.status == 200:
                    data = await response.json()
                    for mod in data.get("data", []):
                        if mod.get("status", {}).get("value") == "loaded":
                            return mod["id"]
        if cached:
            return cached[0]["id"]
    except Exception as e:
        print(f"[JOBS] model pick failed: {e}")
    return None


@app.post("/api/jobs/run")
async def run_job_now(request: Request):
    """Run an on-demand job to completion (inline) and write its output.

    Body: {job: <skill name>, params?: {...}, agent_id?: int}
    Returns the job run record. This endpoint is the future cron hook —
    jobs only run while the app process is up.
    """
    import uuid as _uuid
    from database.job_crud import create_job_run, finish_job_run, get_job_run
    from database.crud import create_conversation, add_message
    from tools.skills_tool import get_skill
    data = await request.json()
    job_name = (data.get("job") or data.get("name") or "").strip()
    params = data.get("params") or {}
    agent_id = data.get("agent_id")

    skill = get_skill(job_name)
    if not skill:
        raise HTTPException(status_code=404, detail=f"Job/skill '{job_name}' not found")

    model = None
    if agent_id is not None:
        async with get_db() as db:
            agent = await get_agent(db, agent_id)
            if agent:
                model = agent.model
    if not model:
        model = await _pick_jobs_model()
    if not model:
        raise HTTPException(status_code=500, detail="No model available for the job run")

    async with get_db() as db:
        conversation = await create_conversation(db, title=f"Job: {job_name}", agent_id=agent_id)
        conversation_id = conversation["id"]
        run = await create_job_run(db, job_name, params=params, conversation_id=conversation_id)
        user_message = (
            f"Run the job '{job_name}' now. Follow the job instructions exactly "
            f"and deliver the output. Job parameters: {json.dumps(params)}"
        )
        await add_message(db, conversation_id, "user", user_message)
        await db.commit()

    request_id = str(_uuid.uuid4())
    assistant_parts = []
    try:
        async for event in _core_stream_handler(request_id, conversation_id, model=model):
            line = event.strip()
            if line.startswith("data: "):
                try:
                    ev = json.loads(line[6:])
                    if ev.get("type") == "content":
                        assistant_parts.append(ev.get("content", ""))
                except Exception:
                    pass
    except Exception as e:
        print(f"[JOBS] run failed: {e}")
        async with get_db() as db:
            await finish_job_run(db, run["id"], "failed", error=str(e))
            await db.commit()
        raise HTTPException(status_code=500, detail=f"Job run failed: {e}")

    assistant_text = "".join(assistant_parts).strip()
    output_path = None
    async with get_db() as db:
        if assistant_text:
            import os as _os
            from settings import OUTPUTS_DIR
            jobs_dir = _os.path.join(OUTPUTS_DIR, "jobs")
            _os.makedirs(jobs_dir, exist_ok=True)
            output_path = _os.path.join(jobs_dir, f"{run['id']}.md")
            try:
                with open(output_path, "w", encoding="utf-8") as f:
                    f.write(f"# Job: {job_name}\n\n{assistant_text}")
            except Exception as e:
                print(f"[JOBS] output write failed: {e}")
                output_path = None
            await finish_job_run(db, run["id"], "completed", output_path=output_path)
        else:
            await finish_job_run(db, run["id"], "failed", error="No assistant output produced")
        await db.commit()
        final = await get_job_run(db, run["id"])
    return {"run": final, "output": assistant_text[:2000]}


# Tool Approval Endpoints
@app.post("/api/tools/{request_id}/approve")
async def approve_tool_request(request_id: str, payload: dict = None):
    """Approve or deny a pending tool execution (e.g. a run_command request).

    The terminal tool yields a `tool_approval_required` SSE event and pauses;
    the frontend calls this endpoint to resolve it.
    """
    from tools.terminal_tool import approval_manager
    payload = payload or {}
    approval_key = payload.get("approval_key") or f"{request_id}:0"
    approved = bool(payload.get("decision", True))
    if not approval_manager.decide(approval_key, approved):
        raise HTTPException(status_code=404, detail="No pending approval found for this request")
    return {"status": "ok", "decision": approved}


# STT Endpoints
@app.post("/api/stt/transcribe")
async def transcribe_audio(request: Request):
    """
    Transcribe audio to text using STT.
    Accepts multipart/form-data with an audio file.
    """
    try:
        form = await request.form()
        audio_file = form.get("audio")
        if not audio_file:
            raise HTTPException(status_code=400, detail="No audio file provided")

        audio_data = await audio_file.read()
        filename = getattr(audio_file, "filename", "recording.webm") or "recording.webm"

        result = await stt_service.transcribe(audio_data, filename=filename)

        if not result.get("success"):
            raise HTTPException(status_code=500, detail=result.get("error", "Transcription failed"))

        return result
    except HTTPException:
        raise
    except Exception as e:
        import traceback; traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"STT error: {str(e)}")


@app.get("/api/stt/status")
async def stt_status():
    """Check STT availability and current config"""
    status = await stt_service.check_availability()
    status["config"] = {
        "engine": stt_service.config.engine,
        "model": stt_service.config.model,
        "language": stt_service.config.language,
    }
    return status


# Backup Management
@app.get("/api/backup/status")
async def get_backup_status():
    """Get database backup status and history"""
    from backend.database.backup import get_backup_status
    return get_backup_status()


@app.post("/api/backup/run")
async def trigger_backup():
    """Trigger a manual database backup"""
    from backend.database.backup import backup_database
    result = backup_database()
    if result.get("success"):
        return {"status": "success", "message": "Backup completed", "file": result["file"], "size": result["size"]}
    else:
        raise HTTPException(status_code=500, detail=result.get("error", "Backup failed"))


@app.post("/api/backup/scheduler/restart")
async def restart_backup_scheduler():
    """Restart the backup scheduler (e.g., after settings change)"""
    await backup_scheduler.restart()
    return {"status": "success", "message": "Backup scheduler restarted"}


# Settings Management
@app.get("/api/settings")
async def get_settings():
    """Get current application settings"""
    return settings_manager.get_settings()


@app.put("/api/settings")
async def update_settings(request: Request):
    """Update application settings"""
    data = await request.json()
    updated_settings = settings_manager.update_settings(data)
    
    # Update LLM client with new settings if URL or model changed
    if 'llama_cpp_base_url' in data or 'llama_cpp_model' in data:
        settings = settings_manager.get_settings()
        llm_client.base_url = settings.get('llama_cpp_base_url', llm_client.base_url)
        llm_client.model = settings.get('llama_cpp_model', llm_client.model)
        print(f"Updated LLM client: base_url={llm_client.base_url}, model={llm_client.model}")

    # Update STT service config if STT settings changed
    stt_keys = {'stt_engine', 'stt_model', 'stt_language', 'stt_openai_api_key'}
    if stt_keys & set(data.keys()):
        global stt_service
        stt_service = STTService(STTConfig.from_settings(updated_settings))
        print(f"Updated STT service: engine={stt_service.config.engine}, model={stt_service.config.model}")

    return updated_settings


@app.get("/api/audio/{filename}")
async def get_audio_file(filename: str):
    """Serve generated TTS audio files"""
    import mimetypes
    audio_path = os.path.join(UPLOAD_DIR, filename)

    if not os.path.exists(audio_path):
        raise HTTPException(status_code=404, detail="Audio file not found")

    mime_type, _ = mimetypes.guess_type(filename)
    media_type = mime_type or "audio/mpeg"

    return FileResponse(
        audio_path,
        media_type=media_type,
        filename=filename
    )


# ─── Tags ─────────────────────────────────────────────────
@app.put("/api/conversations/{conversation_id}/tags")
async def update_tags(conversation_id: str, request: Request):
    """Update tags for a conversation"""
    data = await request.json()
    tags = data.get("tags", [])
    async with get_db() as db:
        result = await update_conversation_tags(db, conversation_id, tags)
        if not result:
            raise HTTPException(status_code=404, detail="Conversation not found")
        return result


@app.put("/api/conversations/{conversation_id}/agent")
async def update_conversation_agent_endpoint(conversation_id: str, request: Request):
    """Update the agent associated with a conversation"""
    data = await request.json()
    agent_id = data.get("agent_id")
    async with get_db() as db:
        result = await update_conversation_agent(db, conversation_id, agent_id)
        if not result:
            raise HTTPException(status_code=404, detail="Conversation not found")
        return result


# ─── Notes ────────────────────────────────────────────────
@app.get("/api/notes")
async def list_notes():
    """Get all notes"""
    async with get_db() as db:
        notes = await get_all_notes(db)
        return {"notes": notes}


@app.get("/api/conversations/{conversation_id}/notes")
async def get_conversation_notes(conversation_id: str):
    """Get notes for a specific conversation"""
    async with get_db() as db:
        notes = await get_notes_for_conversation(db, conversation_id)
        return {"notes": notes}


@app.post("/api/notes")
async def add_note(request: Request):
    """Create a new note"""
    data = await request.json()
    conversation_id = data.get("conversation_id")
    message_id = data.get("message_id")
    content = data.get("content", "")
    source_text = data.get("source_text")
    
    if not conversation_id or not content.strip():
        raise HTTPException(status_code=400, detail="conversation_id and content are required")
    
    async with get_db() as db:
        note = await create_note(db, conversation_id, message_id, content.strip(), source_text)
        return {"note": note}


@app.delete("/api/notes/{note_id}")
async def delete_note_endpoint(note_id: str):
    """Delete a note"""
    async with get_db() as db:
        success = await delete_note(db, note_id)
        if not success:
            raise HTTPException(status_code=404, detail="Note not found")
        return {"status": "success", "message": "Note deleted"}


# ════════════════════════════════════════════════════════════════════════════
# Books — read-aloud reader (PDF / EPUB with TTS + sentence highlight)
# ════════════════════════════════════════════════════════════════════════════
# Five endpoints, all thin: business logic lives in tools/book_service.py and
# database/book_crud.py. The stream endpoint is the only complex one — it
# walks sentences, runs each through tts_service.stream_speech, and persists
# progress so a disconnect mid-read doesn't lose the place.
from database.book_crud import (
    create_book as db_create_book,
    list_books as db_list_books,
    get_book as db_get_book,
    update_book_progress as db_update_book_progress,
    update_book_sentence_progress as db_update_book_sentence_progress,
    soft_delete_book as db_soft_delete_book,
    set_book_sentences as db_set_book_sentences,
)
from tools.book_service import extract as book_extract, derive_title as book_derive_title


@app.post("/api/books/upload")
async def upload_book(file: UploadFile = File(...)):
    """Upload a PDF or EPUB, extract sentences, store as a Book.

    Extraction runs synchronously in the request: a 300-page book is ~1-2s
    on a warm PyPDF2 import, and the user is waiting on a "loading" toast
    anyway. Background-task it later if real-world latencies are bad.
    """
    raw_name = (file.filename or "").strip()
    if not raw_name:
        raise HTTPException(status_code=400, detail="No file provided")

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="File is empty")

    ext = os.path.splitext(raw_name)[1].lower()
    if ext == ".pdf":
        file_type = "pdf"
    elif ext == ".epub":
        file_type = "epub"
    else:
        raise HTTPException(
            status_code=400,
            detail=f"Unsupported file type '{ext}'. Use .pdf or .epub",
        )

    try:
        max_size = int(settings_manager.get_settings().get("max_upload_size", MAX_UPLOAD_SIZE))
    except Exception:
        max_size = MAX_UPLOAD_SIZE
    if len(content) > max_size:
        raise HTTPException(
            status_code=400,
            detail=f"File too large ({len(content)} bytes). Max: {max_size} bytes",
        )

    os.makedirs(UPLOAD_DIR, exist_ok=True)
    stored = f"{uuid.uuid4()}{ext}"
    file_path = os.path.join(UPLOAD_DIR, stored)
    with open(file_path, "wb") as f:
        f.write(content)

    # Extract — failures here must roll back the file write so we don't leak
    # unreadable uploads into the uploads dir.
    try:
        extracted = book_extract(file_path, file_type)
    except Exception as e:
        try:
            os.remove(file_path)
        except OSError:
            pass
        raise HTTPException(status_code=400, detail=f"Extraction failed: {e}")

    title = book_derive_title(raw_name)
    try:
        async with get_db() as db:
            book = await db_create_book(
                db,
                title=title,
                filepath=file_path,
                file_type=file_type,
                sentences=extracted["sentences"],
                page_map=extracted["page_map"],
                size_bytes=len(content),
            )
    except Exception:
        # DB write failed (e.g. table missing pre-migration) — don't leave
        # a 13MB PDF lying in uploads/ to be cleaned up by hand.
        try:
            os.remove(file_path)
        except OSError:
            pass
        raise

    return {"status": "ok", "book": book, "total_sentences": len(extracted["sentences"])}


@app.post("/api/books/from-url")
async def save_book_from_url(body: Dict[str, Any]):
    """Save a webpage to the library: fetch, extract main content, cache
    sentences + a sanitized article-HTML snapshot for the human reading view.
    Body: {url: str}. Returns the existing entry on re-save (deduped by
    normalized URL). Three response shapes:
      saved     — full article saved ({book, total_sentences})
      hub       — section front: no article text, but article links to pick
                  ({title, links: [{title, url}]}, nothing saved)
      link_card — site blocked scraping: metadata entry with the source link
                  ({book}, sentences empty, opens original in reader)
    """
    from tools.web_extract import (
        fetch_url_text, sentences_from_text, normalize_url, domain_of,
        title_from_url, _HubResult,
    )
    from database.book_crud import get_book_by_source_url

    url = normalize_url((body or {}).get("url", ""))
    if not url.startswith(("http://", "https://")):
        raise HTTPException(status_code=400, detail="Provide a valid http(s) URL")
    async with get_db() as db:
        existing = await get_book_by_source_url(db, url)
        if existing:
            return {"status": "ok", "book": existing, "deduped": True,
                    "total_sentences": existing.get("total_sentences", 0)}
    try:
        fetched = await fetch_url_text(url)
    except _HubResult as h:
        return {"status": "hub", "title": h.title, "source_url": url,
                "domain": domain_of(url), "links": h.links}
    except ValueError as e:
        # Blocked (paywall/bot-wall/challenge) — save a link card so the
        # URL is kept with attribution instead of a bare error.
        async with get_db() as db:
            book = await db_create_book(
                db, title=title_from_url(url), filepath="",
                file_type="url", sentences=[], page_map=[], size_bytes=0,
                source_type="url", source_url=url, domain=domain_of(url),
            )
        return {"status": "link_card", "book": book,
                "detail": f"Site blocks scraping ({e}). Saved as link — open the original in the reader."}
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Fetch failed: {e}")
    try:
        extracted = sentences_from_text(fetched["text"], title=fetched["title"])
    except ValueError as e:
        if fetched.get("links"):
            return {"status": "hub", "title": fetched["title"], "source_url": url,
                    "domain": domain_of(url), "links": fetched["links"]}
        raise HTTPException(status_code=422, detail=str(e))

    os.makedirs(UPLOAD_DIR, exist_ok=True)
    snap = f"{uuid.uuid4()}.txt"
    with open(os.path.join(UPLOAD_DIR, snap), "w", encoding="utf-8") as f:
        f.write(fetched["text"])
    html_path = None
    if fetched.get("html", "").strip():
        from tools.web_extract import article_document
        html_path = os.path.join(UPLOAD_DIR, f"{uuid.uuid4()}.html")
        with open(html_path, "w", encoding="utf-8") as f:
            f.write(article_document(extracted["title"], fetched["html"]))
    async with get_db() as db:
        book = await db_create_book(
            db, title=extracted["title"], filepath=os.path.join(UPLOAD_DIR, snap),
            file_type="url", sentences=extracted["sentences"], page_map=extracted["page_map"],
            size_bytes=len(fetched["text"].encode("utf-8")),
            source_type="url", source_url=url, domain=domain_of(url),
            html_path=html_path,
        )
    return {"status": "ok", "book": book, "total_sentences": len(extracted["sentences"])}


@app.post("/api/books/from-text")
async def save_book_from_text(body: Dict[str, Any]):
    """Save pasted/clipboard text to the library. Body: {title?: str,
    text: str, source_url?: str}. Keeps the optional source URL for attribution."""
    from tools.web_extract import sentences_from_text, normalize_url, domain_of

    text = ((body or {}).get("text") or "").strip()
    if len(text) < 20:
        raise HTTPException(status_code=400, detail="Text is too short — paste at least a paragraph")
    try:
        max_size = int(settings_manager.get_settings().get("max_upload_size", MAX_UPLOAD_SIZE))
    except Exception:
        max_size = MAX_UPLOAD_SIZE
    if len(text.encode("utf-8")) > max_size:
        raise HTTPException(status_code=400, detail="Text too large")
    raw_src = ((body or {}).get("source_url") or "").strip()
    source_url = normalize_url(raw_src) if raw_src else None
    try:
        extracted = sentences_from_text(text, title=((body or {}).get("title") or "").strip())
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))

    os.makedirs(UPLOAD_DIR, exist_ok=True)
    snap = f"{uuid.uuid4()}.txt"
    with open(os.path.join(UPLOAD_DIR, snap), "w", encoding="utf-8") as f:
        f.write(text)
    # Pastes get a minimal article view too (paragraphs only, escaped).
    import html as _html
    paras = "".join(f"<p>{_html.escape(p.strip())}</p>" for p in text.split("\n\n") if p.strip())
    html_path = None
    if paras:
        from tools.web_extract import article_document
        html_path = os.path.join(UPLOAD_DIR, f"{uuid.uuid4()}.html")
        with open(html_path, "w", encoding="utf-8") as f:
            f.write(article_document(extracted["title"], paras))
    async with get_db() as db:
        book = await db_create_book(
            db, title=extracted["title"], filepath=os.path.join(UPLOAD_DIR, snap),
            file_type="text", sentences=extracted["sentences"], page_map=extracted["page_map"],
            size_bytes=len(text.encode("utf-8")),
            source_type="text", source_url=source_url,
            domain=domain_of(source_url) if source_url else None,
            html_path=html_path,
        )
    return {"status": "ok", "book": book, "total_sentences": len(extracted["sentences"])}


@app.get("/api/books")
async def list_books_endpoint():
    """All active books, light payload (no sentences JSON)."""
    async with get_db() as db:
        books = await db_list_books(db)
    return {"books": books}


@app.get("/api/books/{book_id}")
async def get_book_endpoint(book_id: str, include_text: bool = True):
    """Single book. include_text=False skips the sentences payload — used by
    the sidebar / library list to avoid shipping ~500KB of JSON."""
    async with get_db() as db:
        book = await db_get_book(db, book_id, include_text=include_text)
    if not book:
        raise HTTPException(status_code=404, detail="Book not found")
    return book


@app.post("/api/books/{book_id}/progress")
async def update_book_progress_endpoint(book_id: str, body: Dict[str, Any]):
    """Lightweight progress update (e.g. after a manual page flip in the
    reader). Body: {current_page?: int, current_sentence_idx?: int}.
    The stream endpoint's after-sentence updates use db_update_book_progress
    directly; this one is for the UI's manual nav.

    Both fields are optional and INDEPENDENT — sending just current_page
    must NOT clobber current_sentence_idx (the reader overlay's
    _persistPage fires on every page change including the first load,
    which used to reset the sentence cursor to 0 and made the library
    card flash "New" even mid-read, and worse, restart the TTS stream
    from sentence 0 every time the user opened the book). Symmetric
    for current_sentence_idx.
    """
    async with get_db() as db:
        book = await db_get_book(db, book_id, include_text=False)
        if not book:
            raise HTTPException(status_code=404, detail="Book not found")

        page = book.get("current_page") or 1
        idx = book.get("current_sentence_idx") or 0

        has_page = body.get("current_page") is not None
        has_idx = body.get("current_sentence_idx") is not None

        if has_page:
            try:
                page = max(1, int(body["current_page"]))
            except (TypeError, ValueError):
                pass
        if has_idx:
            try:
                idx = max(0, int(body["current_sentence_idx"]))
            except (TypeError, ValueError):
                pass
        elif has_page:
            # current_page moved but the caller didn't give us a new
            # sentence index — derive it from the page_map so the
            # library card stays in sync with the page the reader is
            # actually showing. Without this, opening the reader on a
            # TOC jump would set current_page=200 but leave the
            # sentence cursor at the previous page's first sentence,
            # so the next TTS play would jump backwards to that page.
            # We need page_map here but the get_book call above used
            # include_text=False to keep this endpoint cheap; fetch
            # just the page_map via a targeted read.
            from sqlalchemy import select
            from database.models import Book
            r = await db.execute(select(Book.page_map).where(Book.id == book_id))
            pmap = r.scalar_one_or_none() or []
            try:
                pmap_ints = [int(p) for p in pmap]
            except (TypeError, ValueError):
                pmap_ints = []
            if pmap_ints:
                target = page
                derived = 0
                for i, p in enumerate(pmap_ints):
                    if p >= target:
                        derived = i
                        break
                else:
                    derived = len(pmap_ints) - 1
                idx = derived

        await db_update_book_progress(db, book_id, idx, page)
    return {"ok": True}


@app.post("/api/books/{book_id}/sentences")
async def set_book_sentences_endpoint(book_id: str, body: Dict[str, Any]):
    """Replace the cached sentences + page_map on a book. Used by the
    client's PDF.js-based re-extract (server's PyPDF2 fallback is poor
    for some print-typeset PDFs). Body: {sentences: [{text, page,
    char_start}], page_map: [int]}. Both lists must be the same length
    and non-empty.
    """
    sents = body.get("sentences") or []
    pmap = body.get("page_map") or []
    if not sents or not pmap or len(sents) != len(pmap):
        raise HTTPException(status_code=400, detail="sentences and page_map must be same non-empty length")
    async with get_db() as db:
        ok = await db_set_book_sentences(db, book_id, sents, pmap)
        if not ok:
            raise HTTPException(status_code=404, detail="Book not found")
        # Old sentence indices are invalid against the new list. Repoint
        # to the first sentence on the user's saved current_page (or
        # 0 if we have no idea). current_page is preserved as-is.
        book = await db_get_book(db, book_id, include_text=False)
        if book and book.get("current_page"):
            target = int(book["current_page"])
            for i, p in enumerate(pmap):
                if p >= target:
                    await db_update_book_progress(db, book_id, i, target)
                    break
            else:
                await db_update_book_progress(db, book_id, 0, target)
        else:
            await db_update_book_progress(db, book_id, 0, 1)
    return {"ok": True, "total_sentences": len(sents)}


@app.post("/api/books/{book_id}/reextract")
async def reextract_book_endpoint(book_id: str):
    """Re-run the sentence extraction on the book's file (or re-fetch the
    URL for saved pages). Replaces the cached sentences + page_map; resets
    progress to 0.
    """
    from tools.book_service import extract as book_extract
    async with get_db() as db:
        book = await db_get_book(db, book_id, include_text=False)
    if not book:
        raise HTTPException(status_code=404, detail="Book not found")
    try:
        html_snapshot = None
        if book.get("source_type") == "url" and book.get("source_url"):
            from tools.web_extract import fetch_url_text, sentences_from_text, article_document
            fetched = await fetch_url_text(book["source_url"])
            extracted = sentences_from_text(fetched["text"], title=fetched["title"])
            sentences, page_map = extracted["sentences"], extracted["page_map"]
            if fetched.get("html", "").strip():
                html_snapshot = article_document(extracted["title"], fetched["html"])
        elif book.get("file_type") in ("url", "text"):
            from tools.web_extract import sentences_from_text, article_document
            import html as _html_mod
            with open(book["filepath"], "r", encoding="utf-8", errors="ignore") as f:
                text = f.read()
            extracted = sentences_from_text(text, title=book.get("title", ""))
            sentences, page_map = extracted["sentences"], extracted["page_map"]
            paras = "".join(f"<p>{_html_mod.escape(p.strip())}</p>"
                            for p in text.split("\n\n") if p.strip())
            if paras:
                html_snapshot = article_document(extracted["title"], paras)
        else:
            if not os.path.exists(book["filepath"]):
                raise HTTPException(status_code=410, detail="File no longer exists on disk")
            extracted = book_extract(book["filepath"], book["file_type"])
            sentences, page_map = extracted["sentences"], extracted["page_map"]
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Extraction failed: {e}")
    if not sentences:
        raise HTTPException(status_code=422, detail="No sentences extracted")
    html_path = None
    if html_snapshot:
        os.makedirs(UPLOAD_DIR, exist_ok=True)
        html_path = os.path.join(UPLOAD_DIR, f"{uuid.uuid4()}.html")
        with open(html_path, "w", encoding="utf-8") as f:
            f.write(html_snapshot)
    async with get_db() as db:
        await db_set_book_sentences(db, book_id, sentences, page_map, html_path=html_path)
        # Reset progress so the next open starts at the top of the
        # new sentence list. (Old sentence indices are no longer valid.)
        await db_update_book_progress(db, book_id, 0, 1)
    return {"ok": True, "total_sentences": len(sentences)}


@app.delete("/api/books/{book_id}")
async def delete_book_endpoint(book_id: str):
    """Soft-delete a book and remove its file."""
    async with get_db() as db:
        ok = await db_soft_delete_book(db, book_id)
    if not ok:
        raise HTTPException(status_code=404, detail="Book not found")
    return {"status": "ok"}


@app.get("/api/books/{book_id}/file")
async def get_book_file(book_id: str):
    """Serve the original PDF/EPUB for the in-browser viewer (PDF iframe).
    EPUB is served as application/epub+zip; PDFs use the browser's built-in
    viewer for the page-display side of the read-aloud experience."""
    import mimetypes
    async with get_db() as db:
        book = await db_get_book(db, book_id, include_text=False)
    if not book:
        raise HTTPException(status_code=404, detail="Book not found")
    path = book["filepath"]
    if not os.path.exists(path):
        raise HTTPException(status_code=410, detail="File no longer exists on disk")
    mime, _ = mimetypes.guess_type(path)
    return FileResponse(path, media_type=mime or "application/octet-stream")


@app.get("/api/books/{book_id}/stream")
async def stream_book(book_id: str, request: Request, from_idx: int = 0):
    """Stream one NDJSON event per sentence: {idx, page, sentence, seg, url}.

    Client flow:
      1. Open the stream; the first event tells it the current sentence.
      2. Play the segment URL (reuses the TTS cache — re-reads are instant).
      3. On segment end, the next event arrives and the highlight advances.
      4. Progress is persisted server-side per sentence, so a disconnect
         mid-read resumes from `current_sentence_idx` on reopen.
    """
    async with get_db() as db:
        book = await db_get_book(db, book_id, include_text=True)
    if not book:
        raise HTTPException(status_code=404, detail="Book not found")

    sentences = book.get("sentences") or []
    page_map = book.get("page_map") or []
    if not sentences:
        raise HTTPException(status_code=400, detail="Book has no extractable sentences")

    from_idx = max(0, min(from_idx, len(sentences) - 1))

    # Reuse the tool_executor's TTSService — same instance the chat uses,
    # so the TTS cache is shared (re-reads of the same sentence are instant).
    tts_svc = tool_executor.tts_service

    stop_flag = threading.Event()
    watcher = asyncio.create_task(_watch_disconnect(request, stop_flag))

    async def event_stream():
        try:
            for i in range(from_idx, len(sentences)):
                if await request.is_disconnected():
                    break
                sent = sentences[i].get("text", "") if isinstance(sentences[i], dict) else sentences[i]
                page = page_map[i] if i < len(page_map) else (i + 1)
                # Skip noise: empty sentences, or very short ones (often
                # caused by PDFs with spaced-out text like "A L S O" which
                # the splitter can't merge). Without this, a 40k-sentence
                # stream is mostly TTS calls on junk. Threshold 2: keeps
                # "OK" / "I" type valid fragments, drops 1-char artifacts.
                stripped = sent.strip()
                if not stripped or len(stripped) < 2:
                    continue
                # One bad sentence (TTS engine error, OOM, etc.) must not
                # kill the whole 40k-sentence stream — wrap each in its own
                # try/except, yield an error event, and continue.
                try:
                    had_audio = False
                    async for filename, url in tts_svc.stream_speech(
                        text=sent, should_stop=stop_flag.is_set
                    ):
                        had_audio = True
                        yield json.dumps({
                            "idx": i,
                            "page": page,
                            "sentence": sent,
                            "seg": filename,
                            "url": url,
                        }) + "\n"
                    if not had_audio:
                        # TTS service returned nothing (engine disabled, etc.)
                        # — don't advance the highlight; the next valid
                        # sentence will catch the user up.
                        continue
                except Exception as e:
                    print(f"[books] TTS failed at idx={i}: {e}")
                    continue
                # Persist the sentence cursor after each sentence finishes so
                # a disconnect mid-read resumes from there. Only the SENTENCE
                # cursor is updated here — current_page (the page the user is
                # actually looking at) is owned by the reader overlay's
                # progress endpoint, and this stream must never clobber it
                # (an old/slow stream finishing late would otherwise walk
                # the resume page backwards, sentence by sentence).
                try:
                    async with get_db() as db:
                        await db_update_book_sentence_progress(db, book_id, i + 1)
                except Exception as e:
                    print(f"[books] progress persist failed at idx={i}: {e}")
            yield json.dumps({"done": True}) + "\n"
        except Exception as e:
            import traceback
            traceback.print_exc()
            yield json.dumps({"error": str(e)}) + "\n"
        finally:
            watcher.cancel()

    return StreamingResponse(
        event_stream(),
        media_type="application/x-ndjson",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@app.get("/api/books/{book_id}/article")
async def get_book_article(book_id: str):
    """Serve the sanitized article-HTML snapshot for the human reading view
    (url/text saves). The HTML is sanitized at save time (tag/attribute
    allowlist) and rendered client-side in a sandboxed iframe, so embedded
    scripts/trackers cannot run. 404 when the entry has no article snapshot
    (PDF/EPUB uploads, link cards)."""
    async with get_db() as db:
        book = await db_get_book(db, book_id, include_text=False)
        if not book:
            raise HTTPException(status_code=404, detail="Book not found")
        # html_path is intentionally not in the public book dict —
        # resolve the snapshot location with a targeted read.
        from sqlalchemy import select
        from database.models import Book
        r = await db.execute(select(Book.html_path).where(Book.id == book_id))
        path = r.scalar_one_or_none() or ""
    if not path or not os.path.exists(path):
        raise HTTPException(status_code=404, detail="No article view for this entry")
    return FileResponse(path, media_type="text/html")


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host=APP_HOST, port=APP_PORT, debug=DEBUG)
